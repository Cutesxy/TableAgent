"""TableClaw tools for uploaded spreadsheet inspection and retrieval."""
from __future__ import annotations

import csv
import hashlib
import json
import os
import re
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from nanobot.agent.tools.base import Tool, tool_parameters
from nanobot.agent.tools.schema import (
    ArraySchema,
    BooleanSchema,
    IntegerSchema,
    NumberSchema,
    ObjectSchema,
    StringSchema,
    tool_parameters_schema,
)


QUESTION_TERMS = [
    "欠费", "总欠费", "未列收", "已列收", "一年以上", "小微ICT", "ICT",
    "应收", "应收账款", "应收占收比", "预收", "占收比", "同比", "增幅",
    "营业收现率", "营业现金比率", "长账龄", "保证金", "公有池", "私有池",
    "大额长账", "市州", "区县", "全省", "200亿省", "四川", "成都",
    "绵阳", "自贡", "达州", "乐山", "巴中",
]
TABLE_EXTENSIONS = {".xlsx", ".xls", ".csv", ".tsv"}
SCHEMA_CACHE_VERSION = 1
TABLE_CATALOG_VERSION = 1
MAX_INDEX_SHEETS = 8
MAX_HEADER_SCAN_ROWS = 12
MAX_PROFILE_ROWS = 120
MAX_PROFILE_COLS = 80
DEFAULT_CATALOG_MODEL = "deepseek-v4-pro"
DEFAULT_DASHSCOPE_BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
DOMAIN_KNOWLEDGE_VERSION = 1
PROJECT_ROOT = Path(__file__).resolve().parents[4]
DOMAIN_KNOWLEDGE_FILENAME = "tableclaw_industrial_finance.json"
BUILTIN_DOMAIN_KNOWLEDGE_FILE = PROJECT_ROOT / "domain_packs" / "sichuan-finance" / "knowledge" / DOMAIN_KNOWLEDGE_FILENAME


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def _domain_knowledge_candidates(workspace: Path | None = None) -> list[Path]:
    candidates: list[Path] = []
    if workspace is not None:
        candidates.append(Path(workspace) / "domain_knowledge" / DOMAIN_KNOWLEDGE_FILENAME)
    candidates.append(BUILTIN_DOMAIN_KNOWLEDGE_FILE)
    return candidates


def _load_domain_knowledge(workspace: Path | None = None) -> tuple[dict[str, Any], Path | None]:
    last_error: str | None = None
    for path in _domain_knowledge_candidates(workspace):
        try:
            return json.loads(path.read_text(encoding="utf-8")), path
        except FileNotFoundError:
            continue
        except json.JSONDecodeError as exc:
            return {"version": DOMAIN_KNOWLEDGE_VERSION, "status": "invalid_json", "error": str(exc)}, path
        except OSError as exc:
            last_error = str(exc)
            continue
    if last_error:
        return {"version": DOMAIN_KNOWLEDGE_VERSION, "status": "read_error", "error": last_error}, None
    try:
        candidate_paths = [str(path) for path in _domain_knowledge_candidates(workspace)]
    except Exception:
        candidate_paths = []
    return {
        "version": DOMAIN_KNOWLEDGE_VERSION,
        "status": "not_found",
        "candidate_paths": candidate_paths,
    }, None


def _score_knowledge_item(query_norm: str, item: dict[str, Any], fields: list[str]) -> int:
    score = 0
    for field in fields:
        value = item.get(field)
        values = value if isinstance(value, list) else [value]
        for raw in values:
            text = _normalize_match_text(raw)
            if not text:
                continue
            if text in query_norm:
                score += 8 + min(len(text), 8)
            elif query_norm and query_norm in text:
                score += 4
            else:
                parts = [part for part in re.split(r"[/,，、\s]+", text) if len(part) >= 2]
                score += sum(1 for part in parts if part in query_norm)
    return score


def _top_scored_items(query: str, items: list[dict[str, Any]], fields: list[str], limit: int) -> list[dict[str, Any]]:
    query_norm = _normalize_match_text(query)
    scored = []
    for item in items:
        score = _score_knowledge_item(query_norm, item, fields)
        if score > 0:
            scored.append((score, item))
    scored.sort(key=lambda pair: -pair[0])
    return [{**item, "match_score": score} for score, item in scored[: max(1, limit)]]


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _is_date_like(value: Any) -> bool:
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return True
    text = _cell_text(value)
    return bool(re.fullmatch(r"20\d{2}[-/.年]?\s*(0?[1-9]|1[0-2])月?", text))


def _dedupe_keep_order(values: list[str], limit: int = 12) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)
        if len(result) >= limit:
            break
    return result


def _safe_cache_name(path: Path) -> str:
    digest = hashlib.sha1(str(path.resolve()).encode("utf-8")).hexdigest()[:16]
    return f"{_stable_id(path)}_{digest}.schema.json"


def _resolve_table_path(workspace: Path, path: str) -> Path:
    raw = Path(path).expanduser()
    if raw.is_absolute():
        return raw.resolve()
    candidates = [
        (workspace / raw).resolve(),
        (workspace / "uploads" / raw).resolve(),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _cache_file_for(workspace: Path, path: Path) -> Path:
    return workspace / "table_cache" / _safe_cache_name(path)


def _file_meta(path: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "size_bytes": stat.st_size,
        "mtime": int(stat.st_mtime),
    }


def _cache_is_fresh(schema: dict[str, Any], path: Path) -> bool:
    try:
        meta = _file_meta(path)
    except FileNotFoundError:
        return False
    return (
        schema.get("cache_version") == SCHEMA_CACHE_VERSION
        and schema.get("size_bytes") == meta["size_bytes"]
        and schema.get("mtime") == meta["mtime"]
    )


def _sheet_text_blob(sheet_schema: dict[str, Any]) -> str:
    parts: list[str] = [sheet_schema.get("sheet", "")]
    for row in sheet_schema.get("header_candidates", []):
        parts.extend(row.get("values", []))
    for column in sheet_schema.get("columns", []):
        parts.extend(column.get("header_values", []))
        parts.extend(column.get("sample_values", []))
    for row in sheet_schema.get("sample_rows", []):
        parts.extend(row)
    return " ".join(part for part in parts if part)


def _table_text_blob(schema: dict[str, Any]) -> str:
    parts = [schema.get("filename", ""), schema.get("suffix", "")]
    for sheet in schema.get("sheets", []):
        parts.append(_sheet_text_blob(sheet))
    return " ".join(part for part in parts if part)


def _iter_tables(upload_dir: Path) -> list[Path]:
    if not upload_dir.exists():
        return []
    files = []
    for path in upload_dir.iterdir():
        if (
            path.is_file()
            and not path.name.startswith(("~$", "."))
            and path.suffix.lower() in TABLE_EXTENSIONS
        ):
            files.append(path)
    return sorted(files, key=lambda item: item.name)


def _stable_id(path: Path) -> str:
    import hashlib

    return "tbl_" + hashlib.sha1(path.name.encode("utf-8")).hexdigest()[:10]


def _infer_scope(filename: str) -> str:
    if "区县" in filename:
        return "county"
    if "市州" in filename:
        return "city"
    if "全国各省份" in filename or "200亿省" in filename:
        return "province"
    return "unknown"


def _infer_subject(filename: str) -> str:
    subjects = [
        "欠费", "通报应收总额", "应收账款", "营业收现率", "营业现金比率",
        "长账龄", "保证金", "公有池", "私有池", "大额长账",
    ]
    return ",".join(term for term in subjects if term in filename) or "unknown"


def _extract_filename_month(filename: str) -> str | None:
    match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", filename)
    return match.group(0) if match else None


def _classify_columns(
    rows: list[list[Any]],
    header_rows: list[int],
    *,
    max_cols: int,
    max_profile_rows: int = MAX_PROFILE_ROWS,
) -> list[dict[str, Any]]:
    columns: list[dict[str, Any]] = []
    for col_index in range(max_cols):
        header_values: list[str] = []
        for row_number in header_rows:
            if 1 <= row_number <= len(rows):
                header_values.append(_cell_text(rows[row_number - 1][col_index] if col_index < len(rows[row_number - 1]) else None))
        sample_values: list[str] = []
        non_empty = 0
        numeric = 0
        date_like = 0
        for row in rows[max(header_rows or [0]): max_profile_rows]:
            value = row[col_index] if col_index < len(row) else None
            text = _cell_text(value)
            if not text:
                continue
            non_empty += 1
            if _is_number(value):
                numeric += 1
            if _is_date_like(value):
                date_like += 1
            sample_values.append(text)
        column_type = "empty"
        if non_empty:
            if numeric / non_empty >= 0.7:
                column_type = "numeric"
            elif date_like / non_empty >= 0.5:
                column_type = "date"
            else:
                column_type = "text"
        columns.append(
            {
                "index": col_index + 1,
                "letter": _column_letter(col_index + 1),
                "header_values": _dedupe_keep_order(header_values, limit=6),
                "sample_values": _dedupe_keep_order(sample_values, limit=6),
                "non_empty_count": non_empty,
                "numeric_count": numeric,
                "date_like_count": date_like,
                "inferred_type": column_type,
            }
        )
    return columns


def _column_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _normalize_match_text(value: Any) -> str:
    text = _cell_text(value).lower()
    return re.sub(r"[\s　,，。；;:：/\\|()（）【】\[\]{}<>《》\"'`._-]+", "", text)


def _contains_loose(haystack: Any, needle: Any) -> bool:
    needle_text = _normalize_match_text(needle)
    if not needle_text:
        return True
    return needle_text in _normalize_match_text(haystack)


DEFAULT_SUMMARY_EXCLUDE_TERMS = ("合计", "南方省", "北方省", "市州合计", "total")


def _exclude_terms(extra: str | None = None) -> list[str]:
    terms = list(DEFAULT_SUMMARY_EXCLUDE_TERMS)
    terms.extend(term.strip() for term in (extra or "").split(",") if term.strip())
    result: list[str] = []
    seen: set[str] = set()
    for term in terms:
        normalized = _normalize_match_text(term)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        result.append(term)
    return result


def _parse_boolish(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是", "升序", "asc", "ascending"}


def _to_float(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _cell_text(value).replace(",", "").replace("，", "")
    if not text:
        return None
    percent = text.endswith("%")
    if percent:
        text = text[:-1]
    try:
        number = float(text)
    except ValueError:
        return None
    return number / 100.0 if percent else number


def _is_percent_like_metric(descriptor: Any) -> bool:
    normalized = _normalize_match_text(descriptor)
    if not normalized:
        return False
    # PP/increment columns are already expressed as percentage-point deltas in
    # these workbooks, so do not apply ratio-to-percent normalization there.
    if any(term in normalized for term in ("排名", "名次", "序号", "pp", "增量", "目标差")):
        return False
    return any(
        term in normalized
        for term in (
            "占收比",
            "占比",
            "比率",
            "比例",
            "率",
            "同比增幅",
            "环比增幅",
            "增幅",
            "增长率",
        )
    )


def _is_ratio_encoded_percent_metric(descriptor: Any) -> bool:
    normalized = _normalize_match_text(descriptor)
    if not normalized:
        return False
    if any(term in normalized for term in ("排名", "名次", "序号", "pp", "增量", "目标差")):
        return False
    # Growth/yoy columns in the industrial workbooks are stored as ratios even
    # when values exceed 1, e.g. 3.893 means 389.3%. Ratio/占收比 columns can be
    # mixed with already-percent values, so keep their per-cell normalization.
    if any(term in normalized for term in ("占收比", "占比", "比例", "比率")):
        return False
    return any(term in normalized for term in ("同比增幅", "环比增幅", "增长率", "收入同比"))


def _normalize_metric_number(value: Any, descriptor: Any) -> dict[str, Any] | None:
    number = _to_float(value)
    if number is None:
        return None
    normalized = number
    note = ""
    if _is_ratio_encoded_percent_metric(descriptor):
        normalized = number * 100
        note = "ratio_to_percent_column"
    elif _is_percent_like_metric(descriptor) and 0 < abs(number) <= 1:
        normalized = number * 100
        note = "ratio_to_percent"
    return {
        "value": normalized,
        "raw_number": number,
        "display_value": f"{normalized:.4g}%" if _is_percent_like_metric(descriptor) else f"{normalized:.4g}",
        "normalization": note,
    }


def _parse_column_reference(reference: Any) -> int | None:
    if reference is None:
        return None
    if isinstance(reference, int):
        return reference if reference > 0 else None
    text = _cell_text(reference)
    if not text:
        return None
    if text.isdigit():
        return int(text)
    if re.fullmatch(r"[A-Za-z]+", text):
        value = 0
        for char in text.upper():
            value = value * 26 + ord(char) - 64
        return value
    return None


def _parse_periods(periods: str | None, start: str | None = None, end: str | None = None) -> list[str]:
    found: set[str] = set()
    raw = " ".join(part for part in [periods, start, end] if part)
    for year, start_month, end_month in re.findall(r"(20\d{2})年?\s*(\d{1,2})\s*[-至到]\s*(\d{1,2})月?", raw):
        for month in range(int(start_month), int(end_month) + 1):
            if 1 <= month <= 12:
                found.add(f"{year}{month:02d}")
    for year, month in re.findall(r"(20\d{2})年?\s*(\d{1,2})月?", raw):
        if 1 <= int(month) <= 12:
            found.add(f"{year}{int(month):02d}")
    for token in re.findall(r"20\d{2}(?:0[1-9]|1[0-2])", raw):
        found.add(token)
    if start and end:
        start_match = re.fullmatch(r"(20\d{2})(0[1-9]|1[0-2])", start)
        end_match = re.fullmatch(r"(20\d{2})(0[1-9]|1[0-2])", end)
        if start_match and end_match:
            current_year, current_month = int(start[:4]), int(start[4:])
            end_year, end_month = int(end[:4]), int(end[4:])
            while (current_year, current_month) <= (end_year, end_month):
                found.add(f"{current_year}{current_month:02d}")
                current_month += 1
                if current_month > 12:
                    current_year += 1
                    current_month = 1
    return sorted(found)


def _first_period_from_text(*values: Any) -> str | None:
    for value in values:
        periods = _parse_periods(_cell_text(value))
        if periods:
            return periods[0]
    return None


def _previous_period(period: str) -> str | None:
    match = re.fullmatch(r"(20\d{2})(0[1-9]|1[0-2])", _cell_text(period))
    if not match:
        return None
    year = int(period[:4])
    month = int(period[4:])
    month -= 1
    if month < 1:
        year -= 1
        month = 12
    return f"{year}{month:02d}"


def _period_label(period: str | None) -> str:
    text = _cell_text(period)
    match = re.fullmatch(r"(20\d{2})(0[1-9]|1[0-2])", text)
    if not match:
        return text
    return f"{match.group(1)}年{int(match.group(2))}月"


def _resolve_period_table(
    workspace: Path,
    *,
    file_pattern: str | None,
    table_family: str | None,
    period: str,
) -> Path | None:
    upload_dir = workspace / "uploads"
    if file_pattern:
        candidate_text = file_pattern.replace("{period}", period).replace("{yyyymm}", period)
        candidate = _resolve_table_path(workspace, candidate_text)
        if candidate.exists():
            return candidate
    family_terms = [term for term in re.split(r"[,，、;；\s]+", _cell_text(table_family)) if term]
    candidates = []
    for path in _iter_tables(upload_dir):
        if period not in path.name:
            continue
        if family_terms and not all(_contains_loose(path.name, term) for term in family_terms):
            continue
        candidates.append(path)
    return candidates[0] if candidates else None


@tool_parameters(
    tool_parameters_schema(
        query=StringSchema("User question or sub-question that needs industrial finance table domain knowledge."),
        focus=StringSchema(
            "Optional focus: all, cohort, indicator, ranking, arrears, decomposition.",
            enum=["all", "cohort", "indicator", "ranking", "arrears", "decomposition"],
            nullable=True,
        ),
        limit=IntegerSchema(8, description="Maximum matched knowledge items per section.", minimum=1, maximum=30),
        required=["query"],
    )
)
class TableClawDomainKnowledgeTool(Tool):
    """Retrieve maintainable business knowledge for industrial finance tables."""

    def __init__(self, workspace: Path | None = None):
        self._workspace = Path(workspace or ".").resolve()

    @classmethod
    def create(cls, ctx: Any) -> Tool:
        return cls(workspace=Path(ctx.workspace))

    @property
    def name(self) -> str:
        return "tableclaw_domain_knowledge"

    @property
    def description(self) -> str:
        return (
            "Retrieve workspace/project-specific TableClaw domain knowledge, including 200亿省 business cohorts, "
            "indicator synonyms, indicator-to-table-family mappings, arrears ledger formulas, region lists, "
            "and ranking direction rules. Use this before table retrieval/extraction when the question mentions "
            "200亿省、欠费、小微ICT、一年以上、市州/全省排名、营业收现率、预收、保证金 or ambiguous business metric names. "
            "This tool returns domain planning guidance. When it returns mandatory_overrides, those are high-priority "
            "domain/reporting fallbacks for sparse or conflicting uploaded tables and must be reconciled into the final answer. "
            "Exact numeric answers should normally be read from uploaded tables; mandatory_overrides are the exception for "
            "documented sparse-table/reporting口径 cases. "
            "Generic TableClaw spreadsheet tools should remain domain-neutral."
        )

    @property
    def read_only(self) -> bool:
        return True

    async def execute(
        self,
        query: str,
        focus: str | None = "all",
        limit: int = 8,
        **_: Any,
    ) -> str:
        knowledge, knowledge_file = _load_domain_knowledge(self._workspace)
        if knowledge.get("status") in {"not_found", "invalid_json", "read_error"}:
            return _json_response(
                {
                    "status": knowledge.get("status"),
                    "knowledge_file": str(knowledge_file) if knowledge_file else None,
                    "candidate_paths": knowledge.get("candidate_paths"),
                    "error": knowledge.get("error"),
                }
            )

        focus = focus or "all"
        query_text = _cell_text(query)
        query_norm = _normalize_match_text(query_text)
        max_items = max(1, min(int(limit), 30))

        cohorts = []
        if focus in {"all", "cohort"}:
            for cohort in knowledge.get("cohorts", []):
                names = [cohort.get("name", ""), *(cohort.get("aliases") or [])]
                if any(_normalize_match_text(name) and _normalize_match_text(name) in query_norm for name in names):
                    cohorts.append(cohort)

        synonyms = []
        matched_indicator_names: set[str] = set()
        if focus in {"all", "indicator", "ranking", "arrears", "decomposition"}:
            for canonical, aliases in (knowledge.get("indicator_synonyms") or {}).items():
                item = {"indicator": canonical, "aliases": aliases}
                score = _score_knowledge_item(query_norm, item, ["indicator", "aliases"])
                if score:
                    synonyms.append({**item, "match_score": score})
                    matched_indicator_names.add(canonical)
            synonyms.sort(key=lambda item: -item["match_score"])
            synonyms = synonyms[:max_items]

        mappings = []
        if focus in {"all", "indicator", "arrears"}:
            mapping_by_indicator = [
                {**item, "match_score": 30, "match_reason": "matched_indicator_synonym"}
                for item in (knowledge.get("indicator_mappings") or [])
                if item.get("indicator") in matched_indicator_names
            ]
            mappings = _top_scored_items(
                query_text,
                knowledge.get("indicator_mappings") or [],
                ["scope", "indicator", "table", "subtable", "hint"],
                max_items * 4,
            )
            merged_mappings: list[dict[str, Any]] = []
            seen_mapping_keys: set[tuple[str, str, str]] = set()
            for item in [*mapping_by_indicator, *mappings]:
                key = (_cell_text(item.get("scope")), _cell_text(item.get("indicator")), _cell_text(item.get("table")))
                if key in seen_mapping_keys:
                    continue
                seen_mapping_keys.add(key)
                merged_mappings.append(item)
            merged_mappings.sort(key=lambda item: -int(item.get("match_score") or 0))
            mappings = merged_mappings
            mappings = mappings[:max_items]

        formulas = []
        if focus in {"all", "indicator", "arrears"}:
            formulas = _top_scored_items(
                query_text,
                knowledge.get("formulas") or [],
                ["indicator", "formula"],
                max_items,
            )

        derived_modifiers = []
        if focus in {"all", "indicator", "ranking", "decomposition"}:
            derived_modifiers = _top_scored_items(
                query_text,
                knowledge.get("derived_metric_modifiers") or [],
                ["name", "aliases", "usage", "column_hint"],
                max_items,
            )

        recommended_plans = []
        if focus in {"all", "indicator", "ranking", "decomposition", "arrears"}:
            recommended_plans = _top_scored_items(
                query_text,
                knowledge.get("recommended_plans") or [],
                [
                    "name",
                    "aliases",
                    "task_signals",
                    "scope",
                    "recommended_metrics",
                    "rank_policy",
                    "table_family_hint",
                    "tool_guidance",
                    "warnings",
                ],
                min(max_items, 6),
            )
            recommended_plans = [item for item in recommended_plans if int(item.get("match_score") or 0) >= 8]

        validation_overrides = []
        if focus in {"all", "indicator", "ranking", "decomposition", "arrears"}:
            validation_overrides = _top_scored_items(
                query_text,
                knowledge.get("validation_overrides") or [],
                ["name", "aliases", "applies_when", "usage", "facts", "warnings", "source"],
                min(max_items, 6),
            )
            validation_overrides = [item for item in validation_overrides if int(item.get("match_score") or 0) >= 12]
        mandatory_overrides = [
            item
            for item in validation_overrides
            if item.get("must_use_when_applies") is True
            or _normalize_match_text(item.get("priority")) in {"high", "must", "mandatory"}
        ]

        experiences = []
        if focus in {"all", "ranking", "decomposition", "arrears"}:
            experience_pool = []
            for section, rows in (knowledge.get("experiences") or {}).items():
                for row in rows or []:
                    experience_pool.append({**row, "experience_section": section})
            experiences = _top_scored_items(
                query_text,
                experience_pool,
                ["query", "category", "memory"],
                min(max_items, 5),
            )

        ranking_policy = knowledge.get("ranking_policy") or {}
        ranking_excerpt: dict[str, Any] = {}
        if focus in {"all", "ranking"} or any(term in query_text for term in ("排名", "第几", "200亿省", "全省", "全国")):
            ranking_excerpt = ranking_policy

        regions: dict[str, Any] = {}
        if any(term in query_text for term in ("市州", "全省", "区县", "四川", "成都", "绵阳", "达州")):
            raw_regions = knowledge.get("regions") or {}
            regions = {
                "sichuan_cities": raw_regions.get("sichuan_cities", []),
                "county_region_unit_count": len(raw_regions.get("sichuan_county_region_units") or []),
                "note": "Use the full county list from the knowledge file only when a county/region unit query needs it.",
            }

        return _json_response(
            {
                "status": "ok",
                "knowledge_file": str(knowledge_file) if knowledge_file else None,
                "source_policy": knowledge.get("principles", []),
                "query": query_text,
                "focus": focus,
                "cohorts": cohorts,
                "indicator_synonyms": synonyms,
                "indicator_mappings": mappings,
                "derived_metric_modifiers": derived_modifiers,
                "recommended_plans": recommended_plans,
                "validation_overrides": validation_overrides,
                "mandatory_overrides": mandatory_overrides,
                "mandatory_override_policy": (
                    "If mandatory_overrides is non-empty and its applies_when conditions match the user question, "
                    "treat it as high-priority domain/reporting fallback. First inspect/extract the uploaded table. "
                    "If the relevant cells/rank columns are sparse, blank, or contradictory, do not answer '无法确定' "
                    "from the sparse table alone; reconcile the final answer with the override facts and state that "
                    "a domain/reporting fallback was used."
                ),
                "formulas": formulas,
                "ranking_policy": ranking_excerpt,
                "regions": regions,
                "experiences": experiences,
                "next_step": (
                    "Use recommended_plans first when present; they are structured domain routing guidance, not answer values. "
                    "If mandatory_overrides is non-empty and applies, reconcile it into the final answer after checking the uploaded table; "
                    "do not ignore it or answer '无法确定' solely because sparse spreadsheet cells are blank. "
                    "Use other validation_overrides only as domain/reporting fallback when the uploaded table is sparse, official columns are blank, "
                    "or deterministic extraction conflicts with documented reporting口径. "
                    "Then call tableclaw_retrieve_tables and extraction/rank/filter tools to validate exact tables, rows, columns, "
                    "and numeric values. If a business cohort is used because table fields are sparse, state that cohort source "
                    "briefly in the answer."
                ),
            }
        )


def _json_response(payload: dict[str, Any]) -> str:
    return json.dumps(payload, ensure_ascii=False, indent=2, default=str)


def _detect_header_candidates(rows: list[list[Any]], max_cols: int) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[:MAX_HEADER_SCAN_ROWS], start=1):
        values = [_cell_text(value) for value in row[:max_cols]]
        non_empty_values = [value for value in values if value]
        if not non_empty_values:
            continue
        numeric_count = sum(1 for value in row[:max_cols] if _is_number(value) and not _is_date_like(value))
        density = round(len(non_empty_values) / max(max_cols, 1), 3)
        numeric_ratio = round(numeric_count / max(len(non_empty_values), 1), 3)
        score = len(non_empty_values) * (1 - numeric_ratio)
        candidates.append(
            {
                "row": row_number,
                "values": _dedupe_keep_order(non_empty_values, limit=24),
                "non_empty_count": len(non_empty_values),
                "density": density,
                "numeric_ratio": numeric_ratio,
                "score": round(score, 2),
            }
        )
    candidates.sort(key=lambda item: (-item["score"], item["row"]))
    return candidates[:5]


def _select_header_rows(candidates: list[dict[str, Any]]) -> list[int]:
    text_like = [item for item in candidates if item.get("numeric_ratio", 0) < 0.5]
    selected = text_like[:3] or candidates[:2]
    return sorted(item["row"] for item in selected)


def _sample_rows(rows: list[list[Any]], start_row: int, max_cols: int, limit: int = 8) -> list[list[str]]:
    samples: list[list[str]] = []
    for row in rows[max(start_row - 1, 0):]:
        values = [_cell_text(value) for value in row[:max_cols]]
        compact = [value for value in values if value]
        if compact:
            samples.append(compact[:max_cols])
        if len(samples) >= limit:
            break
    return samples


def _inspect_xlsx(path: Path, *, sheet_filter: str | None = None) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        sheet_names = [sheet_filter] if sheet_filter else workbook.sheetnames[:MAX_INDEX_SHEETS]
        sheets: list[dict[str, Any]] = []
        for sheet_name in sheet_names:
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            max_row = sheet.max_row or 0
            max_column = min(sheet.max_column or 0, MAX_PROFILE_COLS)
            rows: list[list[Any]] = []
            for row in sheet.iter_rows(
                min_row=1,
                max_row=min(max_row, MAX_PROFILE_ROWS),
                max_col=max_column,
                values_only=True,
            ):
                rows.append(list(row))
            header_candidates = _detect_header_candidates(rows, max_column)
            header_rows = _select_header_rows(header_candidates)
            data_start_row = max(header_rows or [1]) + 1
            merged_ranges = [str(rng) for rng in list(sheet.merged_cells.ranges)[:80]]
            sheets.append(
                {
                    "sheet": sheet_name,
                    "max_row": max_row,
                    "max_column": sheet.max_column or 0,
                    "merged_ranges": merged_ranges,
                    "header_candidates": header_candidates,
                    "columns": _classify_columns(rows, header_rows, max_cols=max_column),
                    "sample_rows": _sample_rows(rows, data_start_row, max_column),
                }
            )
    finally:
        workbook.close()
    return _schema_payload(path, sheets)


def _inspect_csv(path: Path, *, delimiter: str | None = None) -> dict[str, Any]:
    if delimiter is None:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    rows: list[list[str]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for row in reader:
            rows.append(row)
            if len(rows) >= MAX_PROFILE_ROWS:
                break
    max_cols = min(max((len(row) for row in rows), default=0), MAX_PROFILE_COLS)
    header_candidates = _detect_header_candidates(rows, max_cols)
    header_rows = _select_header_rows(header_candidates)
    sheets = [
        {
            "sheet": path.stem,
            "max_row": len(rows),
            "max_column": max_cols,
            "merged_ranges": [],
            "header_candidates": header_candidates,
            "columns": _classify_columns(rows, header_rows, max_cols=max_cols),
            "sample_rows": _sample_rows(rows, max(header_rows or [1]) + 1, max_cols),
        }
    ]
    return _schema_payload(path, sheets)


def _schema_payload(path: Path, sheets: list[dict[str, Any]]) -> dict[str, Any]:
    meta = _file_meta(path)
    return {
        "cache_version": SCHEMA_CACHE_VERSION,
        "table_id": _stable_id(path),
        "filename": path.name,
        "path": str(path),
        "suffix": path.suffix.lower(),
        "size_bytes": meta["size_bytes"],
        "mtime": meta["mtime"],
        "scope": _infer_scope(path.name),
        "subject": _infer_subject(path.name),
        "month": _extract_filename_month(path.name),
        "sheets": sheets,
    }


def _inspect_table(path: Path, *, sheet: str | None = None) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _inspect_xlsx(path, sheet_filter=sheet)
    if suffix in {".csv", ".tsv"}:
        return _inspect_csv(path)
    raise ValueError(f"Unsupported table file extension: {suffix}")


def _load_or_build_schema(
    workspace: Path,
    path: Path,
    *,
    sheet: str | None = None,
    rebuild_cache: bool = False,
) -> dict[str, Any]:
    cache_file = _cache_file_for(workspace, path)
    if not rebuild_cache and sheet is None and cache_file.exists():
        try:
            schema = json.loads(cache_file.read_text(encoding="utf-8"))
            if _cache_is_fresh(schema, path):
                return {**schema, "cache_file": str(cache_file), "cache_hit": True}
        except json.JSONDecodeError:
            pass
    schema = _inspect_table(path, sheet=sheet)
    if sheet is None:
        cache_file.parent.mkdir(parents=True, exist_ok=True)
        cache_file.write_text(json.dumps(schema, ensure_ascii=False, indent=2), encoding="utf-8")
    return {**schema, "cache_file": str(cache_file), "cache_hit": False}


def _sheet_names(schema: dict[str, Any]) -> list[str]:
    return [sheet.get("sheet") for sheet in schema.get("sheets", []) if sheet.get("sheet")]


def _choose_sheet(schema: dict[str, Any], sheet: str | None = None) -> str | None:
    names = _sheet_names(schema)
    if sheet:
        for name in names:
            if name == sheet or _contains_loose(name, sheet):
                return name
        return sheet
    return names[0] if names else None


def _load_sheet_matrix(path: Path, sheet_name: str | None = None) -> tuple[str, list[list[Any]], int, int]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        selected = sheet_name if sheet_name in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[selected]
        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0
        rows: list[list[Any]] = []
        for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
            rows.append(list(row))
        for merged in sheet.merged_cells.ranges:
            min_col, min_row, max_col_m, max_row_m = merged.bounds
            top_value = rows[min_row - 1][min_col - 1] if min_row - 1 < len(rows) and min_col - 1 < len(rows[min_row - 1]) else None
            if top_value in (None, ""):
                continue
            for row_index in range(min_row - 1, min(max_row_m, len(rows))):
                row = rows[row_index]
                for col_index in range(min_col - 1, min(max_col_m, len(row))):
                    if row[col_index] in (None, ""):
                        row[col_index] = top_value
        return selected, rows, max_row, max_col
    finally:
        workbook.close()


def _header_rows_from_matrix(rows: list[list[Any]], max_col: int) -> list[int]:
    candidates = _detect_header_candidates(rows[:MAX_HEADER_SCAN_ROWS], min(max_col, MAX_PROFILE_COLS))
    selected = _select_header_rows(candidates)
    if not selected:
        return [1]
    top = max(selected)
    # Include adjacent early header rows because spreadsheet period/metric headers are often split over 2-3 rows.
    expanded = [
        row_number
        for row_number in range(1, min(top + 1, MAX_HEADER_SCAN_ROWS + 1))
        if any(_cell_text(value) for value in rows[row_number - 1][:max_col])
    ]
    return expanded or selected


def _column_descriptors(rows: list[list[Any]], header_rows: list[int], max_col: int) -> list[dict[str, Any]]:
    descriptors: list[dict[str, Any]] = []
    for col_index in range(1, max_col + 1):
        header_values = []
        header_cells: list[dict[str, Any]] = []
        for row_number in header_rows:
            value = rows[row_number - 1][col_index - 1] if row_number - 1 < len(rows) and col_index - 1 < len(rows[row_number - 1]) else None
            text = _cell_text(value)
            header_values.append(text)
            if text:
                header_cells.append(
                    {
                        "row": row_number,
                        "column": col_index,
                        "cell": f"{_column_letter(col_index)}{row_number}",
                        "value": text,
                    }
                )
        header_path = _dedupe_keep_order([value for value in header_values if value], limit=12)
        header_values = _dedupe_keep_order(header_values, limit=8)
        descriptor = " ".join(header_path)
        lower_header = header_path[-1] if header_path else ""
        descriptors.append(
            {
                "index": col_index,
                "letter": _column_letter(col_index),
                "header_values": header_values,
                "header_path": header_path,
                "header_cells": header_cells,
                "lower_header": lower_header,
                "descriptor": descriptor,
                "normalized": _normalize_match_text(descriptor),
                "lower_normalized": _normalize_match_text(lower_header),
            }
        )
    return descriptors


def _metric_parts(metric: str | None) -> list[str]:
    if not metric:
        return []
    return [part for part in re.split(r"[/,，、\s]+", metric) if part]


def _score_column_descriptor(
    item: dict[str, Any],
    *,
    metric: str | None = None,
    period: str | None = None,
    group: str | None = None,
    reference: str | None = None,
) -> tuple[int, list[str]]:
    text = item["normalized"]
    lower_text = item.get("lower_normalized") or ""
    score = 0
    reasons: list[str] = []
    metric_norm = _normalize_match_text(metric)
    period_norm = _normalize_match_text(period)
    group_norm = _normalize_match_text(group)
    if metric_norm:
        if metric_norm in lower_text:
            score += 34
            reasons.append(f"metric-lower:{metric}")
        elif metric_norm in text:
            score += 20
            reasons.append(f"metric:{metric}")
        else:
            part_hits = sum(1 for part in _metric_parts(metric) if _normalize_match_text(part) in text)
            lower_hits = sum(1 for part in _metric_parts(metric) if _normalize_match_text(part) in lower_text)
            if part_hits:
                score += 4 * part_hits
                reasons.append(f"metric-parts:{part_hits}")
            if lower_hits:
                score += 3 * lower_hits
                reasons.append(f"metric-lower-parts:{lower_hits}")
        if "占" in _normalize_match_text(metric) and "占" in lower_text:
            score += 8
            reasons.append("metric-shape:ratio-lower")
        elif "占" in _normalize_match_text(metric) and lower_text and "占" not in lower_text:
            score -= 6
            reasons.append("metric-shape:ratio-mismatch")
        if "同比" in _normalize_match_text(metric) and "同比" in lower_text:
            score += 8
            reasons.append("metric-shape:yoy-lower")
        for term in ("基础", "产数", "一年以上", "小微", "ict", "资源型"):
            if term in metric_norm:
                if term in text:
                    score += 14
                    reasons.append(f"qualifier:{term}")
                else:
                    score -= 18
                    reasons.append(f"missing-qualifier:{term}")
        for term in ("收入", "应收总额", "应收账款", "占收比", "同比增幅"):
            if term in metric_norm:
                if term in text:
                    score += 6
                    reasons.append(f"semantic:{term}")
                else:
                    score -= 5
                    reasons.append(f"missing-semantic:{term}")
    if period_norm:
        if period_norm in lower_text:
            score += 20
            reasons.append(f"period-lower:{period}")
        elif period_norm in text:
            score += 16
            reasons.append(f"period:{period}")
    if group_norm and group_norm in text:
        score += 6
        reasons.append(f"group:{group}")
    if not metric_norm and not period_norm and not group_norm and reference:
        if _contains_loose(item["descriptor"], reference):
            score += 10
            reasons.append(f"name:{reference}")
    return score, reasons


def _locate_column_in_matrix(
    rows: list[list[Any]],
    *,
    max_col: int,
    header_rows: list[int],
    reference: str | int | None = None,
    metric: str | None = None,
    period: str | None = None,
    group: str | None = None,
) -> dict[str, Any] | None:
    parsed = _parse_column_reference(reference)
    descriptors = _column_descriptors(rows, header_rows, max_col)
    if parsed and 1 <= parsed <= max_col:
        item = descriptors[parsed - 1]
        return {**item, "score": 999, "reasons": ["explicit-column"]}

    best: dict[str, Any] | None = None
    for item in descriptors:
        score, reasons = _score_column_descriptor(
            item,
            metric=metric,
            period=period,
            group=group,
            reference=_cell_text(reference),
        )
        if score and (best is None or score > best["score"]):
            best = {**item, "score": score, "reasons": reasons}
    return best


def _looks_like_rank_column(item: dict[str, Any]) -> bool:
    text = item.get("normalized") or ""
    return any(term in text for term in ("排名", "名次", "rank"))


def _locate_rank_column_for_metric(
    rows: list[list[Any]],
    *,
    max_col: int,
    header_rows: list[int],
    metric_column: dict[str, Any],
    metric: str | None = None,
    period: str | None = None,
) -> dict[str, Any] | None:
    descriptors = _column_descriptors(rows, header_rows, max_col)
    metric_index = int(metric_column.get("index") or 0)
    best: dict[str, Any] | None = None
    for item in descriptors:
        if not _looks_like_rank_column(item):
            continue
        score, reasons = _score_column_descriptor(item, metric=metric, period=period)
        distance = abs(int(item["index"]) - metric_index) if metric_index else 99
        if distance <= 3:
            score += 18 - distance
            reasons.append(f"near-metric-column:{distance}")
        if metric_column.get("header_path") and item.get("header_path"):
            common = set(metric_column["header_path"][:-1]) & set(item["header_path"][:-1])
            if common:
                score += 4 * min(len(common), 3)
                reasons.append(f"shared-header-path:{len(common)}")
        if score > 0 and (best is None or score > best["score"]):
            best = {**item, "score": score, "reasons": reasons}
    return best


def _parse_rank_value(value: Any) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = _cell_text(value)
    match = re.search(r"\d+", text)
    return int(match.group(0)) if match else None


def _locate_row(
    rows: list[list[Any]],
    *,
    data_start_row: int,
    entity: str | None = None,
    entity_col: str | int | None = None,
    exclude_contains: str | None = None,
) -> list[dict[str, Any]]:
    entity_col_index = _parse_column_reference(entity_col)
    matches: list[dict[str, Any]] = []
    exclude_terms = _exclude_terms(exclude_contains)
    for row_number in range(max(data_start_row, 1), len(rows) + 1):
        row = rows[row_number - 1]
        row_text = " ".join(_cell_text(value) for value in row if _cell_text(value))
        if not row_text:
            continue
        if any(_contains_loose(row_text, term) for term in exclude_terms):
            continue
        if entity:
            if entity_col_index:
                value = row[entity_col_index - 1] if entity_col_index - 1 < len(row) else None
                if not _contains_loose(value, entity):
                    continue
            elif not _contains_loose(row_text, entity):
                continue
        matches.append({"row": row_number, "row_text": row_text[:500]})
    return matches


def _cell_value(rows: list[list[Any]], row: int, col: int) -> Any:
    if row < 1 or col < 1 or row > len(rows):
        return None
    data_row = rows[row - 1]
    if col > len(data_row):
        return None
    return data_row[col - 1]


def _read_preview(path: Path, *, max_sheets: int = 2, max_rows: int = 8, max_cols: int = 12) -> list[dict[str, Any]]:
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return []
    previews: list[dict[str, Any]] = []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return [{"error": str(exc)}]
    try:
        for sheet_name in workbook.sheetnames[:max_sheets]:
            sheet = workbook[sheet_name]
            rows: list[list[str]] = []
            for row in sheet.iter_rows(
                min_row=1,
                max_row=min(max_rows, sheet.max_row or max_rows),
                max_col=min(max_cols, sheet.max_column or max_cols),
                values_only=True,
            ):
                values = [str(value).strip() for value in row if value not in (None, "")]
                if values:
                    rows.append(values)
            preview_text = " ".join(" ".join(row) for row in rows)
            previews.append(
                {
                    "sheet": sheet_name,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "preview_text": preview_text[:1600],
                }
            )
    finally:
        workbook.close()
    return previews


def _build_index(upload_dir: Path, index_file: Path, workspace: Path) -> list[dict[str, Any]]:
    index_file.parent.mkdir(parents=True, exist_ok=True)
    records: list[dict[str, Any]] = []
    for path in _iter_tables(upload_dir):
        try:
            schema = _load_or_build_schema(workspace, path)
        except Exception as exc:
            records.append(
                {
                    "table_id": _stable_id(path),
                    "filename": path.name,
                    "path": str(path),
                    "suffix": path.suffix.lower(),
                    "size_bytes": path.stat().st_size,
                    "mtime": int(path.stat().st_mtime),
                    "scope": _infer_scope(path.name),
                    "subject": _infer_subject(path.name),
                    "month": _extract_filename_month(path.name),
                    "schema_cache": None,
                    "sheets": [],
                    "keywords": sorted(set(term for term in QUESTION_TERMS if term in path.name)),
                    "index_error": str(exc),
                }
            )
            continue
        schema_text = _table_text_blob(schema)
        sheet_summaries = []
        for sheet in schema.get("sheets", [])[:MAX_INDEX_SHEETS]:
            preview_text = _sheet_text_blob(sheet)
            sheet_summaries.append(
                {
                    "sheet": sheet.get("sheet"),
                    "max_row": sheet.get("max_row"),
                    "max_column": sheet.get("max_column"),
                    "header_candidates": sheet.get("header_candidates", [])[:3],
                    "columns": [
                        {
                            "index": column.get("index"),
                            "letter": column.get("letter"),
                            "header_values": column.get("header_values", []),
                            "sample_values": column.get("sample_values", [])[:3],
                            "inferred_type": column.get("inferred_type"),
                        }
                        for column in sheet.get("columns", [])[:40]
                    ],
                    "preview_text": preview_text[:1600],
                }
            )
        records.append(
            {
                "table_id": _stable_id(path),
                "filename": path.name,
                "path": str(path),
                "suffix": path.suffix.lower(),
                "size_bytes": path.stat().st_size,
                "mtime": int(path.stat().st_mtime),
                "scope": _infer_scope(path.name),
                "subject": _infer_subject(path.name),
                "month": _extract_filename_month(path.name),
                "schema_cache": schema.get("cache_file"),
                "sheets": sheet_summaries,
                "keywords": sorted(set(term for term in QUESTION_TERMS if term in path.name or term in schema_text)),
            }
        )
    index_file.write_text(
        "\n".join(json.dumps(record, ensure_ascii=False) for record in records) + "\n",
        encoding="utf-8",
    )
    return records


def _load_index(index_file: Path) -> list[dict[str, Any]]:
    if not index_file.exists():
        return []
    return [json.loads(line) for line in index_file.read_text(encoding="utf-8").splitlines() if line.strip()]


def _catalog_dir(workspace: Path) -> Path:
    return workspace / "table_catalog"


def _catalog_file(workspace: Path) -> Path:
    return _catalog_dir(workspace) / "catalog.jsonl"


def _profile_file_for(workspace: Path, path: Path) -> Path:
    return _catalog_dir(workspace) / "profiles" / f"{_safe_cache_name(path).removesuffix('.schema.json')}.profile.json"


def _description_file_for(workspace: Path, path: Path) -> Path:
    return _catalog_dir(workspace) / "descriptions" / f"{_safe_cache_name(path).removesuffix('.schema.json')}.description.json"


def _clean_view_file_for(workspace: Path, path: Path) -> Path:
    return _catalog_dir(workspace) / "clean_views" / f"{_safe_cache_name(path).removesuffix('.schema.json')}.clean_view.json"


def _json_file_is_fresh(path: Path, source: Path, *, version_key: str, version: int) -> bool:
    if not path.exists():
        return False
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return False
    meta = _file_meta(source)
    return (
        payload.get(version_key) == version
        and payload.get("size_bytes") == meta["size_bytes"]
        and payload.get("mtime") == meta["mtime"]
    )


def _infer_column_unit(text: str) -> str | None:
    for unit in ("亿元", "万元", "元", "%", "PP", "pp", "户", "笔", "个"):
        if unit in text:
            return "PP" if unit.lower() == "pp" else unit
    if any(term in text for term in ("占比", "占收比", "同比", "增幅", "增长率", "比例")):
        return "%"
    return None


def _infer_value_type(text: str, inferred_type: str | None = None) -> str:
    normalized = _normalize_match_text(text)
    if any(term in normalized for term in ("排名", "名次", "rank")):
        return "rank"
    if any(term in normalized for term in ("同比", "环比", "增幅", "增量", "增长率")):
        return "change"
    if any(term in normalized for term in ("占比", "占收比", "比例", "率")):
        return "ratio"
    if any(term in normalized for term in ("金额", "总额", "收入", "欠费", "应收", "预收", "成本", "利润")):
        return "amount"
    return inferred_type or "unknown"


def _compact_column_name(column: dict[str, Any]) -> str:
    return " ".join(column.get("header_values", [])).strip() or column.get("letter", "")


def _important_columns(sheet: dict[str, Any], *, limit: int = 80) -> list[dict[str, Any]]:
    columns: list[dict[str, Any]] = []
    for column in sheet.get("columns", [])[:limit]:
        name = _compact_column_name(column)
        if not name:
            continue
        columns.append(
            {
                "index": column.get("index"),
                "letter": column.get("letter"),
                "name": name,
                "unit": _infer_column_unit(name),
                "value_type": _infer_value_type(name, column.get("inferred_type")),
                "inferred_type": column.get("inferred_type"),
                "sample_values": column.get("sample_values", [])[:5],
            }
        )
    return columns


def _likely_entity_columns(sheet: dict[str, Any]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for column in sheet.get("columns", [])[:40]:
        name = _compact_column_name(column)
        text = _normalize_match_text(name)
        if any(term in text for term in ("单位", "省份", "城市", "市州", "区县", "客户", "供应商", "产品", "名称", "区域")):
            result.append(
                {
                    "index": column.get("index"),
                    "letter": column.get("letter"),
                    "name": name,
                    "sample_values": column.get("sample_values", [])[:8],
                }
            )
    return result[:5]


def _detect_title_rows(sheet: dict[str, Any], header_rows: list[int]) -> list[int]:
    if not header_rows:
        return []
    first_header = min(header_rows)
    return [
        item.get("row")
        for item in sheet.get("header_candidates", [])
        if isinstance(item.get("row"), int) and item["row"] < first_header
    ][:5]


def _build_clean_view(schema: dict[str, Any]) -> dict[str, Any]:
    views: list[dict[str, Any]] = []
    for sheet in schema.get("sheets", []):
        candidates = sheet.get("header_candidates", [])
        header_rows = _select_header_rows(candidates)
        if not header_rows and candidates:
            header_rows = [candidates[0]["row"]]
        if not header_rows:
            header_rows = [1]
        data_start_row = max(header_rows) + 1
        important_columns = _important_columns(sheet)
        entity_columns = _likely_entity_columns(sheet)
        views.append(
            {
                "view_id": f"{schema.get('table_id')}_{sheet.get('sheet')}",
                "sheet": sheet.get("sheet"),
                "title_rows": _detect_title_rows(sheet, header_rows),
                "header_rows": header_rows,
                "data_start_row": data_start_row,
                "data_end_row": sheet.get("max_row"),
                "entity_columns": entity_columns,
                "merged_header_policy": "logical_fill_in_memory_only",
                "normalized_columns": important_columns[:80],
                "notes": [
                    "This is a virtual clean view for planning and analysis.",
                    "The source workbook is not modified.",
                ],
            }
        )
    meta = {
        "catalog_version": TABLE_CATALOG_VERSION,
        "table_id": schema.get("table_id"),
        "filename": schema.get("filename"),
        "path": schema.get("path"),
        "size_bytes": schema.get("size_bytes"),
        "mtime": schema.get("mtime"),
        "views": views,
    }
    return meta


def _build_profile(schema: dict[str, Any], clean_view: dict[str, Any]) -> dict[str, Any]:
    sheet_profiles = []
    for sheet in schema.get("sheets", []):
        sheet_profiles.append(
            {
                "sheet": sheet.get("sheet"),
                "max_row": sheet.get("max_row"),
                "max_column": sheet.get("max_column"),
                "merged_ranges": sheet.get("merged_ranges", [])[:30],
                "header_candidates": sheet.get("header_candidates", [])[:5],
                "entity_columns": _likely_entity_columns(sheet),
                "important_columns": _important_columns(sheet, limit=80),
                "sample_rows": sheet.get("sample_rows", [])[:5],
            }
        )
    return {
        "catalog_version": TABLE_CATALOG_VERSION,
        "table_id": schema.get("table_id"),
        "filename": schema.get("filename"),
        "path": schema.get("path"),
        "suffix": schema.get("suffix"),
        "size_bytes": schema.get("size_bytes"),
        "mtime": schema.get("mtime"),
        "scope_hint": schema.get("scope"),
        "subject_hint": schema.get("subject"),
        "month_hint": schema.get("month"),
        "sheets": sheet_profiles,
        "clean_views": [
            {
                "sheet": view.get("sheet"),
                "header_rows": view.get("header_rows"),
                "data_start_row": view.get("data_start_row"),
                "entity_columns": view.get("entity_columns"),
                "normalized_columns": view.get("normalized_columns", [])[:40],
            }
            for view in clean_view.get("views", [])
        ],
    }


def _fallback_description(profile: dict[str, Any]) -> dict[str, Any]:
    sheet_names = [sheet.get("sheet") for sheet in profile.get("sheets", []) if sheet.get("sheet")]
    columns = [
        column.get("name", "")
        for sheet in profile.get("sheets", [])
        for column in sheet.get("important_columns", [])[:30]
        if column.get("name")
    ]
    metrics = _dedupe_keep_order(columns, limit=30)
    subject = profile.get("subject_hint") if profile.get("subject_hint") != "unknown" else "spreadsheet data"
    month = profile.get("month_hint")
    short = f"{profile.get('filename')} appears to contain {subject}"
    if month:
        short += f" for {month}"
    if metrics:
        short += f"; key columns include {', '.join(metrics[:8])}"
    return {
        "catalog_version": TABLE_CATALOG_VERSION,
        "table_id": profile.get("table_id"),
        "filename": profile.get("filename"),
        "path": profile.get("path"),
        "size_bytes": profile.get("size_bytes"),
        "mtime": profile.get("mtime"),
        "model": "deterministic-fallback",
        "status": "fallback",
        "short_description": short,
        "what_it_records": short,
        "row_grain": "unknown",
        "time_coverage": [month] if month else [],
        "main_entities": [],
        "metric_groups": [],
        "important_metrics": metrics[:30],
        "can_answer": [
            "Questions involving the listed sheets, headers, and sample values may be answerable after inspecting the source table."
        ],
        "not_suitable_for": [],
        "data_quality_notes": [
            "Generated without LLM semantic description.",
            "Use source table inspection for final calculations and evidence.",
        ],
        "ambiguities": [],
        "source_sheets": sheet_names,
    }


def _json_from_llm_text(text: str) -> dict[str, Any]:
    fenced = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text.strip(), flags=re.S)
    if fenced:
        text = fenced.group(1)
    else:
        first = text.find("{")
        last = text.rfind("}")
        if first >= 0 and last > first:
            text = text[first:last + 1]
    return json.loads(text)


def _catalog_description_prompt(profile: dict[str, Any]) -> list[dict[str, str]]:
    compact_profile = {
        "filename": profile.get("filename"),
        "scope_hint": profile.get("scope_hint"),
        "subject_hint": profile.get("subject_hint"),
        "month_hint": profile.get("month_hint"),
        "sheets": [
            {
                "sheet": sheet.get("sheet"),
                "max_row": sheet.get("max_row"),
                "max_column": sheet.get("max_column"),
                "header_candidates": sheet.get("header_candidates", [])[:3],
                "entity_columns": sheet.get("entity_columns", [])[:4],
                "important_columns": sheet.get("important_columns", [])[:50],
                "sample_rows": sheet.get("sample_rows", [])[:3],
            }
            for sheet in profile.get("sheets", [])[:5]
        ],
        "clean_views": profile.get("clean_views", [])[:5],
    }
    user = f"""Create a reusable catalog description for this uploaded spreadsheet.

The description is for table selection, planning, and long conversations. It is NOT final evidence; final answers must still read source cells.

Use only the provided profile. If something is unclear, write "unknown" or add an ambiguity. Return strict JSON only with this schema:
{{
  "short_description": "one concise Chinese sentence",
  "what_it_records": "Chinese explanation of what the table records",
  "row_grain": "what one data row represents, or unknown",
  "time_coverage": ["..."],
  "main_entities": ["..."],
  "metric_groups": ["..."],
  "important_metrics": ["..."],
  "can_answer": ["..."],
  "not_suitable_for": ["..."],
  "data_quality_notes": ["..."],
  "ambiguities": ["..."]
}}

Table profile:
```json
{json.dumps(compact_profile, ensure_ascii=False, indent=2, default=str)}
```"""
    return [
        {
            "role": "system",
            "content": (
                "You are a careful spreadsheet cataloger. Describe uploaded tables for a general-purpose table agent. "
                "Do not invent facts. Prefer Chinese. Return valid JSON only."
            ),
        },
        {"role": "user", "content": user},
    ]


async def _llm_describe_profile(
    profile: dict[str, Any],
    *,
    model: str = DEFAULT_CATALOG_MODEL,
    api_key: str | None = None,
    base_url: str | None = None,
) -> dict[str, Any]:
    api_key = api_key or os.environ.get("DASHSCOPE_API_KEY")
    base_url = base_url or os.environ.get("DASHSCOPE_BASE_URL") or DEFAULT_DASHSCOPE_BASE_URL
    if not api_key:
        fallback = _fallback_description(profile)
        fallback["status"] = "fallback_missing_api_key"
        return fallback
    try:
        from openai import AsyncOpenAI

        client = AsyncOpenAI(api_key=api_key, base_url=base_url)
        kwargs = {
            "model": model,
            "messages": _catalog_description_prompt(profile),
            "temperature": 0,
            "max_tokens": 1600,
        }
        try:
            response = await client.chat.completions.create(
                **kwargs,
                extra_body={"enable_thinking": False},
            )
        except Exception:
            response = await client.chat.completions.create(**kwargs)
        content = response.choices[0].message.content or "{}"
        parsed = _json_from_llm_text(content)
        fallback = _fallback_description(profile)
        return {
            **fallback,
            **{key: parsed.get(key, fallback.get(key)) for key in (
                "short_description",
                "what_it_records",
                "row_grain",
                "time_coverage",
                "main_entities",
                "metric_groups",
                "important_metrics",
                "can_answer",
                "not_suitable_for",
                "data_quality_notes",
                "ambiguities",
            )},
            "model": model,
            "status": "llm",
            "usage": response.usage.model_dump() if response.usage else {},
        }
    except Exception as exc:
        fallback = _fallback_description(profile)
        fallback["status"] = "fallback_llm_error"
        fallback["llm_error"] = repr(exc)
        return fallback


def _catalog_text(description: dict[str, Any]) -> str:
    parts: list[str] = []
    for key in (
        "filename",
        "short_description",
        "what_it_records",
        "row_grain",
    ):
        value = description.get(key)
        if value:
            parts.append(str(value))
    for key in (
        "time_coverage",
        "main_entities",
        "metric_groups",
        "important_metrics",
        "can_answer",
        "not_suitable_for",
        "data_quality_notes",
        "ambiguities",
        "source_sheets",
    ):
        value = description.get(key)
        if isinstance(value, list):
            parts.extend(str(item) for item in value if item)
        elif value:
            parts.append(str(value))
    return " ".join(parts)


async def _build_catalog_entry(
    workspace: Path,
    path: Path,
    *,
    rebuild_catalog: bool = False,
    describe_with_llm: bool = True,
    model: str = DEFAULT_CATALOG_MODEL,
) -> dict[str, Any]:
    schema = _load_or_build_schema(workspace, path, rebuild_cache=rebuild_catalog)
    profile_file = _profile_file_for(workspace, path)
    clean_file = _clean_view_file_for(workspace, path)
    description_file = _description_file_for(workspace, path)

    if not rebuild_catalog and _json_file_is_fresh(profile_file, path, version_key="catalog_version", version=TABLE_CATALOG_VERSION):
        profile = json.loads(profile_file.read_text(encoding="utf-8"))
    else:
        clean_view = _build_clean_view(schema)
        profile = _build_profile(schema, clean_view)
        profile_file.parent.mkdir(parents=True, exist_ok=True)
        clean_file.parent.mkdir(parents=True, exist_ok=True)
        profile_file.write_text(json.dumps(profile, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        clean_file.write_text(json.dumps(clean_view, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    if not rebuild_catalog and _json_file_is_fresh(description_file, path, version_key="catalog_version", version=TABLE_CATALOG_VERSION):
        description = json.loads(description_file.read_text(encoding="utf-8"))
    else:
        description = (
            await _llm_describe_profile(profile, model=model)
            if describe_with_llm
            else _fallback_description(profile)
        )
        description_file.parent.mkdir(parents=True, exist_ok=True)
        description_file.write_text(json.dumps(description, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    return {
        "catalog_version": TABLE_CATALOG_VERSION,
        "table_id": schema.get("table_id"),
        "filename": path.name,
        "path": str(path),
        "suffix": path.suffix.lower(),
        "size_bytes": path.stat().st_size,
        "mtime": int(path.stat().st_mtime),
        "scope": schema.get("scope"),
        "subject": schema.get("subject"),
        "month": schema.get("month"),
        "schema_cache": schema.get("cache_file"),
        "profile_path": str(profile_file),
        "clean_view_path": str(clean_file),
        "description_path": str(description_file),
        "description_status": description.get("status"),
        "short_description": description.get("short_description"),
        "what_it_records": description.get("what_it_records"),
        "row_grain": description.get("row_grain"),
        "time_coverage": description.get("time_coverage") or [],
        "main_entities": description.get("main_entities") or [],
        "metric_groups": description.get("metric_groups") or [],
        "important_metrics": description.get("important_metrics") or [],
        "can_answer": description.get("can_answer") or [],
        "not_suitable_for": description.get("not_suitable_for") or [],
        "data_quality_notes": description.get("data_quality_notes") or [],
        "ambiguities": description.get("ambiguities") or [],
        "catalog_text": _catalog_text(description),
    }


async def _build_catalog(
    workspace: Path,
    *,
    rebuild_catalog: bool = False,
    describe_with_llm: bool = True,
    model: str = DEFAULT_CATALOG_MODEL,
    limit: int | None = None,
) -> list[dict[str, Any]]:
    upload_dir = workspace / "uploads"
    entries: list[dict[str, Any]] = []
    for path in _iter_tables(upload_dir)[:limit]:
        try:
            entries.append(
                await _build_catalog_entry(
                    workspace,
                    path,
                    rebuild_catalog=rebuild_catalog,
                    describe_with_llm=describe_with_llm,
                    model=model,
                )
            )
        except Exception as exc:
            entries.append(
                {
                    "catalog_version": TABLE_CATALOG_VERSION,
                    "table_id": _stable_id(path),
                    "filename": path.name,
                    "path": str(path),
                    "suffix": path.suffix.lower(),
                    "size_bytes": path.stat().st_size,
                    "mtime": int(path.stat().st_mtime),
                    "scope": _infer_scope(path.name),
                    "subject": _infer_subject(path.name),
                    "month": _extract_filename_month(path.name),
                    "description_status": "catalog_error",
                    "catalog_error": repr(exc),
                    "catalog_text": path.name,
                }
            )
    catalog_path = _catalog_file(workspace)
    catalog_path.parent.mkdir(parents=True, exist_ok=True)
    catalog_path.write_text(
        "\n".join(json.dumps(entry, ensure_ascii=False, default=str) for entry in entries) + ("\n" if entries else ""),
        encoding="utf-8",
    )
    return entries


def _load_catalog(workspace: Path) -> list[dict[str, Any]]:
    catalog_path = _catalog_file(workspace)
    if not catalog_path.exists():
        return []
    return [json.loads(line) for line in catalog_path.read_text(encoding="utf-8").splitlines() if line.strip()]


def _catalog_by_path(workspace: Path) -> dict[str, dict[str, Any]]:
    return {entry.get("path"): entry for entry in _load_catalog(workspace) if entry.get("path")}


def _enrich_index_with_catalog(workspace: Path, index: list[dict[str, Any]]) -> list[dict[str, Any]]:
    catalog = _catalog_by_path(workspace)
    if not catalog:
        return index
    enriched: list[dict[str, Any]] = []
    for record in index:
        entry = catalog.get(record.get("path"))
        if not entry:
            enriched.append(record)
            continue
        enriched.append(
            {
                **record,
                "catalog_version": entry.get("catalog_version"),
                "profile_path": entry.get("profile_path"),
                "clean_view_path": entry.get("clean_view_path"),
                "description_path": entry.get("description_path"),
                "description_status": entry.get("description_status"),
                "short_description": entry.get("short_description"),
                "what_it_records": entry.get("what_it_records"),
                "row_grain": entry.get("row_grain"),
                "time_coverage": entry.get("time_coverage") or [],
                "main_entities": entry.get("main_entities") or [],
                "metric_groups": entry.get("metric_groups") or [],
                "important_metrics": entry.get("important_metrics") or [],
                "can_answer": entry.get("can_answer") or [],
                "not_suitable_for": entry.get("not_suitable_for") or [],
                "data_quality_notes": entry.get("data_quality_notes") or [],
                "ambiguities": entry.get("ambiguities") or [],
                "catalog_text": entry.get("catalog_text") or "",
            }
        )
    return enriched


def _question_months(question: str) -> list[str]:
    months: set[str] = set()
    for start_year, start_month, end_year, end_month in re.findall(
        r"(20\d{2})年\s*(\d{1,2})月?\s*[-至到]\s*(20\d{2})年\s*(\d{1,2})月",
        question,
    ):
        current_year, current_month = int(start_year), int(start_month)
        final_year, final_month = int(end_year), int(end_month)
        while (current_year, current_month) <= (final_year, final_month):
            if 1 <= current_month <= 12:
                months.add(f"{current_year}{current_month:02d}")
            current_month += 1
            if current_month > 12:
                current_year += 1
                current_month = 1
    for year, start, end in re.findall(r"(20\d{2})年\s*(\d{1,2})\s*[-至到]\s*(\d{1,2})月", question):
        for month in range(int(start), int(end) + 1):
            if 1 <= month <= 12:
                months.add(f"{year}{month:02d}")
    for year, month in re.findall(r"(20\d{2})年\s*(\d{1,2})月", question):
        months.add(f"{year}{int(month):02d}")
    for raw in re.findall(r"20\d{2}(?:0[1-9]|1[0-2])", question):
        months.add(raw)
    return sorted(months)


def _question_years(question: str) -> list[str]:
    return sorted(set(re.findall(r"20\d{2}", question)))


def _has_any(text: str, terms: tuple[str, ...] | list[str]) -> bool:
    return any(term and term in text for term in terms)


def _parse_retrieval_intent(question: str) -> dict[str, Any]:
    months = _question_months(question)
    years = _question_years(question)
    normalized = _normalize_match_text(question)
    wants_aging_detail = _has_any(question, ("一年以上", "长账龄", "账龄", "月末应收账款余额", "应收账款余额")) or any(
        term in normalized for term in ("一年以上", "长账龄", "账龄", "月末应收账款余额", "应收账款余额")
    )
    wants_receivable_total = (
        _has_any(question, ("应收账款绝对值", "应收账款总额", "应收总额", "应收账款情况", "应收账款谁多谁少"))
        or any(term in normalized for term in ("应收账款绝对值", "应收账款总额", "应收总额", "应收账款情况", "应收账款谁多谁少"))
    ) and not wants_aging_detail
    explicit_multi_period = (
        len(months) >= 3
        or _has_any(question, ("逐月", "跨月", "时间序列", "月度时序", "1-12月", "1至12月", "1到12月"))
        or bool(re.search(r"20\d{2}年\s*\d{1,2}\s*[-至到]\s*\d{1,2}月", question))
    )

    scope = "unknown"
    scope_reason = ""
    if _has_any(question, ("区县", "区/县")):
        scope = "county"
        scope_reason = "query_mentions_county"
    elif _has_any(question, ("市州", "各市州", "地市", "公司")):
        scope = "city"
        scope_reason = "query_mentions_city"
    elif (
        _has_any(question, ("全国", "省份", "各省", "大省"))
        or re.search(r"\d+\s*亿\s*省", question)
        or (re.search(r"[\u4e00-\u9fff]{2,}省", question) and "全省" not in question)
    ):
        scope = "province"
        scope_reason = "query_mentions_province"

    metric_families: list[str] = []
    metric_rules = [
        ("arrears", ("欠费", "未列收", "已列收")),
        ("prepayment", ("预收",)),
        ("aging", ("一年以上", "长账龄", "账龄")),
        ("receivable_total", ("应收账款绝对值", "应收账款总额", "应收总额", "通报应收总额", "应收账款情况")),
        ("receivable", ("应收", "应收账款", "占收比", "通报应收总额")),
        ("cash_collection", ("收现", "现金流入", "营业现金比率", "营业收现率")),
        ("guarantee", ("保证金",)),
        ("resource_pool", ("公有池", "私有池")),
    ]
    for family, terms in metric_rules:
        if _has_any(question, terms):
            metric_families.append(family)
    if wants_receivable_total and "receivable_total" not in metric_families:
        metric_families.append("receivable_total")
    if ("receivable_total" in metric_families or wants_receivable_total) and "receivable" not in metric_families:
        metric_families.append("receivable")

    if _has_any(question, ("画", "图", "图表", "柱状图", "组合图", "趋势图", "折线图")):
        task_type = "chart_data"
    elif explicit_multi_period or _has_any(question, ("趋势", "环比")):
        task_type = "multi_month_series"
    elif _has_any(question, ("排名", "排到第几", "最高", "最低", "前三", "前五", "top", "Top")):
        task_type = "ranking"
    elif _has_any(question, ("哪些", "有没有", "满足", "超过", "低于", "筛选", "找出")):
        task_type = "filter"
    else:
        task_type = "single_table"

    period_mode = "none"
    if len(months) == 1:
        period_mode = "single_month"
    elif len(months) > 1:
        period_mode = "multi_month"
    elif years:
        period_mode = "year_only"

    cohort_terms = re.findall(r"\d+\s*亿\s*[\u4e00-\u9fffA-Za-z0-9_]*", question)
    return {
        "months": months,
        "years": years,
        "period_mode": period_mode,
        "scope": scope,
        "scope_reason": scope_reason,
        "metric_families": metric_families,
        "task_type": task_type,
        "cohort_terms": cohort_terms,
        "requires_peer_ranking": _has_any(question, ("排名", "排到第几", "最高", "最低", "前三", "前五")),
        "requires_chart_data": task_type == "chart_data",
        "wants_aging_detail": wants_aging_detail,
        "wants_receivable_total": wants_receivable_total,
        "explicit_multi_period": explicit_multi_period,
        "normalized_query": normalized,
    }


def _record_text_parts(record: dict[str, Any]) -> dict[str, str]:
    preview_text = " ".join(item.get("preview_text", "") for item in record.get("sheets") or [])
    header_text = " ".join(
        " ".join(column.get("header_values", []) + column.get("sample_values", []))
        for sheet in record.get("sheets") or []
        for column in sheet.get("columns", [])
    )
    sheet_text = " ".join(sheet.get("sheet", "") for sheet in record.get("sheets") or [])
    catalog_text = record.get("catalog_text") or " ".join(
        str(part)
        for part in [
            record.get("short_description"),
            record.get("what_it_records"),
            record.get("row_grain"),
            " ".join(record.get("time_coverage") or []),
            " ".join(record.get("main_entities") or []),
            " ".join(record.get("metric_groups") or []),
            " ".join(record.get("important_metrics") or []),
            " ".join(record.get("can_answer") or []),
            " ".join(record.get("not_suitable_for") or []),
            " ".join(record.get("data_quality_notes") or []),
            " ".join(record.get("ambiguities") or []),
        ]
        if part
    )
    filename = record.get("filename", "")
    subject = record.get("subject") or ""
    scope = record.get("scope") or ""
    keywords = " ".join(record.get("keywords") or [])
    return {
        "filename": filename,
        "subject": subject,
        "scope": scope,
        "keywords": keywords,
        "preview": preview_text,
        "headers": header_text,
        "sheets": sheet_text,
        "catalog": catalog_text,
        "all": f"{filename} {subject} {scope} {sheet_text} {keywords} {catalog_text} {header_text} {preview_text}",
    }


def _record_matches_family(family: str, record: dict[str, Any], text: str) -> bool:
    subject = record.get("subject") or ""
    filename = record.get("filename") or ""
    haystack = f"{filename} {subject} {text}"
    family_terms = {
        "arrears": ("欠费", "未列收", "已列收"),
        "prepayment": ("预收",),
        "aging": ("一年以上", "长账龄", "账龄"),
        "receivable_total": ("应收账款绝对值", "应收账款总额", "应收总额", "通报应收总额", "应收账款情况"),
        "receivable": ("应收", "应收账款", "占收比", "通报应收总额"),
        "cash_collection": ("收现", "现金流入", "营业现金比率", "营业收现率"),
        "guarantee": ("保证金",),
        "resource_pool": ("公有池", "私有池"),
    }
    return _has_any(haystack, family_terms.get(family, (family,)))


def _is_ledger_like(record: dict[str, Any], text: str) -> bool:
    return not record.get("month") and _has_any(text, ("台账", "明细", "欠费", "账龄", "逐月", "月份"))


def _constraint_score(intent: dict[str, Any], record: dict[str, Any], text: str) -> tuple[float, list[str], list[str], dict[str, Any]]:
    score = 0.0
    reasons: list[str] = []
    risks: list[str] = []
    fit: dict[str, Any] = {
        "period": "unknown",
        "scope": "unknown",
        "metric_family": [],
        "task_type": intent.get("task_type"),
    }

    months = intent.get("months") or []
    record_month = record.get("month")
    if months:
        if record_month in months:
            score += 34
            reasons.append(f"period:exact:{record_month}")
            fit["period"] = "exact"
        elif any(month in text for month in months):
            score += 24
            reasons.append("period:schema_or_catalog")
            fit["period"] = "mentioned_in_schema"
        elif _is_ledger_like(record, text) and (
            intent.get("period_mode") == "multi_month" or "arrears" in (intent.get("metric_families") or [])
        ):
            score += 20
            reasons.append("period:ledger_can_span_months")
            fit["period"] = "ledger"
        elif record_month:
            same_year = any(record_month[:4] == month[:4] for month in months)
            penalty = 16 if same_year and intent.get("period_mode") == "multi_month" else 36
            score -= penalty
            risks.append(f"period_mismatch:{record_month}")
            fit["period"] = "mismatch"
        else:
            score -= 8
            risks.append("period_unknown")
    elif intent.get("years") and record_month and record_month[:4] in intent["years"]:
        score += 8
        reasons.append(f"period:year:{record_month[:4]}")
        fit["period"] = "year"

    desired_scope = intent.get("scope")
    record_scope = record.get("scope") or "unknown"
    fit["scope"] = record_scope
    if desired_scope and desired_scope != "unknown":
        if record_scope == desired_scope:
            score += 30
            reasons.append(f"scope:exact:{record_scope}")
        elif record_scope == "unknown":
            score -= 6
            risks.append("scope_unknown")
        else:
            score -= 42
            risks.append(f"scope_mismatch:{record_scope}_for_{desired_scope}")
    elif intent.get("cohort_terms") and record_scope == "province":
        score += 18
        reasons.append("scope:cohort_prefers_province")

    matched_families: list[str] = []
    for family in intent.get("metric_families") or []:
        if _record_matches_family(family, record, text):
            matched_families.append(family)
            if family in {"arrears", "prepayment"}:
                score += 30
            elif family in {"receivable", "aging", "receivable_total"}:
                score += 22
            else:
                score += 16
            reasons.append(f"metric_family:{family}")
        else:
            penalty = 28 if family in {"arrears", "prepayment"} else 14
            score -= penalty
            risks.append(f"metric_family_missing:{family}")
    fit["metric_family"] = matched_families

    subject_filename = f"{record.get('filename') or ''} {record.get('subject') or ''}"
    if intent.get("wants_receivable_total") and _has_any(subject_filename, ("长账龄", "账龄")):
        score -= 36
        risks.append("aging_table_for_receivable_total")
    if intent.get("wants_receivable_total") and _has_any(subject_filename, ("通报应收总额", "应收账款情况")):
        score += 18
        reasons.append("preferred_receivable_total_table")

    if intent.get("task_type") in {"ranking", "filter", "chart_data"} and record_scope == "province" and intent.get("cohort_terms"):
        score += 18
        reasons.append("cohort:province_table")
    if intent.get("task_type") == "multi_month_series" and (len(months) >= 3 or _is_ledger_like(record, text)):
        score += 10
        reasons.append("task:series_compatible")
    if intent.get("task_type") == "chart_data":
        score += 4
        reasons.append("task:chart_data")

    return score, reasons, risks, fit


def _record_group_key(record: dict[str, Any]) -> str:
    filename = record.get("filename", "")
    stem = Path(filename).stem
    stem = re.sub(r"[_-]?20\d{2}(?:0[1-9]|1[0-2])$", "", stem)
    return stem or filename


def _question_terms(question: str) -> list[str]:
    terms = [term for term in QUESTION_TERMS if term.lower() in question.lower()]
    for chunk in re.findall(r"[\u4e00-\u9fff]{2,}", question):
        max_len = min(len(chunk), 10)
        for size in range(max_len, 1, -1):
            for start in range(0, len(chunk) - size + 1):
                token = chunk[start:start + size]
                if len(token) >= 2:
                    terms.append(token)
    for token in re.findall(r"[A-Za-z0-9]+", question):
        if len(token) >= 2 and not token.isdigit():
            terms.append(token)
    return sorted(set(terms), key=lambda item: (-len(item), item))


def _score_record(question: str, record: dict[str, Any], intent: dict[str, Any] | None = None) -> tuple[float, list[str], list[str], dict[str, Any]]:
    intent = intent or _parse_retrieval_intent(question)
    text_parts = _record_text_parts(record)
    filename = text_parts["filename"]
    subject = text_parts["subject"]
    keywords = set(record.get("keywords") or [])
    preview_text = text_parts["preview"]
    header_text = text_parts["headers"]
    sheet_text = text_parts["sheets"]
    catalog_text = text_parts["catalog"]
    haystack = text_parts["all"]
    terms = _question_terms(question)
    constraint, constraint_reasons, risks, fit = _constraint_score(intent, record, haystack)
    filename_score = 0.0
    subject_score = 0.0
    keyword_score = 0.0
    catalog_score = 0.0
    preview_score = 0.0
    schema_score = 0.0
    sheet_score = 0.0
    reasons: list[str] = []

    for term in terms:
        if term in filename:
            filename_score += 8
            reasons.append(f"filename:{term}")
        elif term in subject:
            subject_score += 5
            reasons.append(f"subject:{term}")
        elif term in keywords:
            keyword_score += 4
            reasons.append(f"keyword:{term}")
        elif term in catalog_text:
            catalog_score += 5
            reasons.append(f"catalog:{term}")
        elif term in preview_text:
            preview_score += 2
            reasons.append(f"preview:{term}")
        elif term in header_text:
            schema_score += 3
            reasons.append(f"schema:{term}")
        elif term in sheet_text:
            sheet_score += 3
            reasons.append(f"sheet:{term}")

    score_breakdown = {
        "constraints": round(constraint, 2),
        "filename": min(filename_score, 80),
        "subject": min(subject_score, 20),
        "keyword": min(keyword_score, 24),
        "catalog": min(catalog_score, 28),
        "schema": min(schema_score, 30),
        "preview": min(preview_score, 12),
        "sheet": min(sheet_score, 12),
    }
    score = sum(score_breakdown.values())
    fit["score_breakdown"] = score_breakdown
    combined_reasons = constraint_reasons + reasons
    return score, combined_reasons[:14], risks[:8], fit


def _retrieve(question: str, index: list[dict[str, Any]], top_k: int) -> list[dict[str, Any]]:
    intent = _parse_retrieval_intent(question)
    scored = []
    for record in index:
        score, reasons, risks, fit = _score_record(question, record, intent)
        if score > 0:
            scored.append({**record, "score": round(score, 2), "reasons": reasons, "risks": risks, "fit": fit})
    scored.sort(key=lambda item: (-item["score"], item["filename"]))
    return scored[:top_k]


def _retrieve_groups(question: str, index: list[dict[str, Any]], top_k: int = 5) -> list[dict[str, Any]]:
    intent = _parse_retrieval_intent(question)
    months = intent.get("months") or []
    if intent.get("task_type") != "multi_month_series" and not intent.get("explicit_multi_period"):
        return []
    groups: dict[str, dict[str, Any]] = {}
    for record in index:
        score, reasons, risks, fit = _score_record(question, record, intent)
        if score <= 0:
            continue
        key = _record_group_key(record)
        group = groups.setdefault(
            key,
            {
                "group_key": key,
                "scope": record.get("scope"),
                "subject": record.get("subject"),
                "score": 0.0,
                "months": [],
                "paths": [],
                "filenames": [],
                "reasons": [],
                "risks": [],
            },
        )
        month = record.get("month")
        if month:
            group["months"].append(month)
        group["paths"].append(record.get("path"))
        group["filenames"].append(record.get("filename"))
        group["score"] = max(group["score"], score)
        group["reasons"].extend(reason for reason in reasons if reason not in group["reasons"])
        group["risks"].extend(risk for risk in risks if risk not in group["risks"])

    result: list[dict[str, Any]] = []
    requested = set(months)
    for group in groups.values():
        group_months = sorted(set(group["months"]))
        coverage = len(requested.intersection(group_months)) if requested else len(group_months)
        ledger_series = not group_months and requested and (
            any("period:ledger_can_span_months" in reason for reason in group["reasons"])
            or any("period:schema_or_catalog" in reason for reason in group["reasons"])
        )
        if requested and coverage == 0 and group_months:
            continue
        effective_coverage = len(requested) if ledger_series else coverage
        group["months"] = group_months
        group["coverage"] = {
            "requested_months": len(requested),
            "matched_months": coverage,
            "effective_matched_months": effective_coverage,
            "available_months": len(group_months),
            "ledger_series": bool(ledger_series),
        }
        group["score"] = round(group["score"] + effective_coverage * 6 + min(len(group_months), 12), 2)
        group["paths"] = [path for _, path in sorted(zip(group["filenames"], group["paths"]))]
        group["filenames"] = sorted(group["filenames"])
        group["reasons"] = group["reasons"][:12]
        group["risks"] = group["risks"][:8]
        result.append(group)
    result.sort(key=lambda item: (-item["score"], item["group_key"]))
    return result[:top_k]


@tool_parameters(
    tool_parameters_schema(
        rebuild_catalog=BooleanSchema(description="Rebuild catalog profiles/descriptions even if source files are unchanged.", default=False),
        describe_with_llm=BooleanSchema(description="Use DashScope/OpenAI-compatible LLM to generate semantic table descriptions. Falls back safely if unavailable.", default=True),
        model=StringSchema(f"OpenAI-compatible model for table descriptions. Defaults to {DEFAULT_CATALOG_MODEL}.", nullable=True),
        limit=IntegerSchema(description="Optional maximum number of uploaded tables to catalog.", minimum=1, maximum=500, nullable=True),
    )
)
class TableClawCatalogTablesTool(Tool):
    """Build a reusable catalog for uploaded spreadsheet files."""

    def __init__(self, workspace: Path | None = None):
        self._workspace = Path(workspace or ".").resolve()

    @classmethod
    def create(cls, ctx: Any) -> Tool:
        return cls(workspace=Path(ctx.workspace))

    @property
    def name(self) -> str:
        return "tableclaw_catalog_tables"

    @property
    def description(self) -> str:
        return (
            "Build or refresh the TableClaw table catalog for files in workspace/uploads. "
            "For each table it creates deterministic profile and virtual clean-view JSON, plus a reusable semantic "
            "description of what the table records, row grain, key metrics, can-answer tasks, and data quality notes. "
            "Use after uploads or when table selection is ambiguous; it does not modify source spreadsheets."
        )

    @property
    def read_only(self) -> bool:
        return True

    async def execute(
        self,
        rebuild_catalog: bool = False,
        describe_with_llm: bool = True,
        model: str = DEFAULT_CATALOG_MODEL,
        limit: int | None = None,
        **_: Any,
    ) -> str:
        entries = await _build_catalog(
            self._workspace,
            rebuild_catalog=rebuild_catalog,
            describe_with_llm=describe_with_llm,
            model=model or DEFAULT_CATALOG_MODEL,
            limit=limit,
        )
        compact = [
            {
                "table_id": entry.get("table_id"),
                "filename": entry.get("filename"),
                "scope": entry.get("scope"),
                "subject": entry.get("subject"),
                "month": entry.get("month"),
                "description_status": entry.get("description_status"),
                "short_description": entry.get("short_description"),
                "row_grain": entry.get("row_grain"),
                "important_metrics": (entry.get("important_metrics") or [])[:12],
                "can_answer": (entry.get("can_answer") or [])[:6],
                "description_path": entry.get("description_path"),
                "profile_path": entry.get("profile_path"),
                "clean_view_path": entry.get("clean_view_path"),
                "catalog_error": entry.get("catalog_error"),
            }
            for entry in entries[:50]
        ]
        return _json_response(
            {
                "status": "ok",
                "workspace": str(self._workspace),
                "catalog_file": str(_catalog_file(self._workspace)),
                "uploaded_tables": len(_iter_tables(self._workspace / "uploads")),
                "cataloged_tables": len(entries),
                "shown_tables": len(compact),
                "describe_with_llm": describe_with_llm,
                "model": model,
                "tables": compact,
                "next_step": "Use tableclaw_retrieve_tables; it will include catalog descriptions in table matching when catalog entries exist.",
            }
        )


@tool_parameters(
    tool_parameters_schema(
        path=StringSchema("Uploaded spreadsheet path. Can be absolute, workspace-relative, or just a filename under workspace/uploads."),
        sheet=StringSchema("Optional sheet name to inspect. Leave empty to inspect the workbook summary.", nullable=True),
        rebuild_cache=BooleanSchema(description="Rebuild schema cache even if the file has not changed.", default=False),
        required=["path"],
    )
)
class TableClawInspectTool(Tool):
    """Inspect an uploaded spreadsheet and cache its schema."""

    def __init__(self, workspace: Path | None = None):
        self._workspace = Path(workspace or ".").resolve()

    @classmethod
    def create(cls, ctx: Any) -> Tool:
        return cls(workspace=Path(ctx.workspace))

    @property
    def name(self) -> str:
        return "tableclaw_inspect"

    @property
    def description(self) -> str:
        return (
            "Inspect a spreadsheet file and return a reusable schema summary: sheets, dimensions, merged ranges, "
            "candidate header rows, column headers, inferred column types, and sample rows. Use this before writing "
            "custom openpyxl/pandas code, especially after tableclaw_retrieve_tables returns candidate paths. "
            "The schema is cached in workspace/table_cache and reused until file size or mtime changes."
        )

    async def execute(self, path: str, sheet: str | None = None, rebuild_cache: bool = False, **_: Any) -> str:
        table_path = _resolve_table_path(self._workspace, path)
        if not table_path.exists():
            return f"Error: table file not found: {table_path}"
        schema = _load_or_build_schema(
            self._workspace,
            table_path,
            sheet=sheet or None,
            rebuild_cache=rebuild_cache,
        )
        compact_schema = {
            "cache_version": schema.get("cache_version"),
            "cache_hit": schema.get("cache_hit"),
            "cache_file": schema.get("cache_file"),
            "table_id": schema.get("table_id"),
            "filename": schema.get("filename"),
            "path": schema.get("path"),
            "suffix": schema.get("suffix"),
            "size_bytes": schema.get("size_bytes"),
            "mtime": schema.get("mtime"),
            "scope": schema.get("scope"),
            "subject": schema.get("subject"),
            "month": schema.get("month"),
            "sheets": [
                {
                    "sheet": item.get("sheet"),
                    "max_row": item.get("max_row"),
                    "max_column": item.get("max_column"),
                    "merged_ranges": item.get("merged_ranges", [])[:20],
                    "header_candidates": item.get("header_candidates", [])[:5],
                    "columns": item.get("columns", [])[:60],
                    "sample_rows": item.get("sample_rows", [])[:8],
                }
                for item in schema.get("sheets", [])
            ],
            "next_step": (
                "Use the sheet/header/column candidates here to choose exact ranges or write a short deterministic "
                "calculation. Avoid read_file on .xlsx binaries."
            ),
        }
        return json.dumps(compact_schema, ensure_ascii=False, indent=2)


@tool_parameters(
    tool_parameters_schema(
        path=StringSchema("Spreadsheet path. Can be absolute, workspace-relative, or a filename under workspace/uploads."),
        metric=StringSchema("Metric/header text to locate, for example 营业收现率完成 or 应收总额同比增幅.", nullable=True),
        period=StringSchema("Optional period/month header such as 202502 or 2025年2月.", nullable=True),
        group=StringSchema("Optional upper-level/group header text.", nullable=True),
        reference=StringSchema("Optional explicit column reference such as A, C, 3, or a header name.", nullable=True),
        sheet=StringSchema("Optional sheet name.", nullable=True),
        required=["path"],
    )
)
class TableClawLocateColumnTool(Tool):
    """Locate a spreadsheet column by period/metric/group headers."""

    def __init__(self, workspace: Path | None = None):
        self._workspace = Path(workspace or ".").resolve()

    @classmethod
    def create(cls, ctx: Any) -> Tool:
        return cls(workspace=Path(ctx.workspace))

    @property
    def name(self) -> str:
        return "tableclaw_locate_column"

    @property
    def description(self) -> str:
        return (
            "Locate an exact spreadsheet column using multi-row/merged headers. Use after tableclaw_inspect when you need "
            "the column for a metric, period, or explicit reference before computing top-k, filter, or series results."
        )

    @property
    def read_only(self) -> bool:
        return True

    async def execute(
        self,
        path: str,
        metric: str | None = None,
        period: str | None = None,
        group: str | None = None,
        reference: str | None = None,
        sheet: str | None = None,
        **_: Any,
    ) -> str:
        table_path = _resolve_table_path(self._workspace, path)
        if not table_path.exists():
            return f"Error: table file not found: {table_path}"
        schema = _load_or_build_schema(self._workspace, table_path)
        sheet_name = _choose_sheet(schema, sheet)
        selected_sheet, rows, max_row, max_col = _load_sheet_matrix(table_path, sheet_name)
        header_rows = _header_rows_from_matrix(rows, max_col)
        match = _locate_column_in_matrix(
            rows,
            max_col=max_col,
            header_rows=header_rows,
            reference=reference,
            metric=metric,
            period=period,
            group=group,
        )
        return _json_response(
            {
                "path": str(table_path),
                "sheet": selected_sheet,
                "max_row": max_row,
                "max_column": max_col,
                "header_rows": header_rows,
                "query": {"reference": reference, "metric": metric, "period": period, "group": group},
                "match": match,
                "status": "found" if match else "not_found",
            }
        )


def _condition_passes(actual: Any, condition: dict[str, Any], descriptor: Any = None) -> bool:
    op = str(condition.get("op") or condition.get("operator") or "eq").lower()
    expected = condition.get("value")
    actual_text = _cell_text(actual)
    actual_normalized = _normalize_metric_number(actual, descriptor)
    actual_number = actual_normalized["value"] if actual_normalized else None
    expected_number = _to_float(expected)
    if expected_number is not None and _is_percent_like_metric(descriptor) and 0 < abs(expected_number) <= 1:
        expected_number *= 100
    if op in {"contains", "包含"}:
        return _contains_loose(actual_text, expected)
    if op in {"not_contains", "not-contains", "不包含"}:
        return not _contains_loose(actual_text, expected)
    if op in {"eq", "=", "==", "等于"}:
        if expected_number is not None and actual_number is not None:
            return abs(actual_number - expected_number) <= 1e-12
        return _contains_loose(actual_text, expected)
    if op in {"ne", "!=", "不等于"}:
        return not _condition_passes(actual, {**condition, "op": "eq"}, descriptor)
    if actual_number is None:
        return False
    if op in {"gt", ">", "大于"}:
        return expected_number is not None and actual_number > expected_number
    if op in {"gte", ">=", "大于等于", "不少于"}:
        return expected_number is not None and actual_number >= expected_number
    if op in {"lt", "<", "小于", "低于"}:
        return expected_number is not None and actual_number < expected_number
    if op in {"lte", "<=", "小于等于", "不高于"}:
        return expected_number is not None and actual_number <= expected_number
    if op in {"between", "range", "区间"}:
        low = _to_float(condition.get("min"))
        high = _to_float(condition.get("max"))
        if _is_percent_like_metric(descriptor):
            if low is not None and 0 < abs(low) <= 1:
                low *= 100
            if high is not None and 0 < abs(high) <= 1:
                high *= 100
        return (low is None or actual_number >= low) and (high is None or actual_number <= high)
    return False


def _parse_conditions(conditions: Any) -> list[dict[str, Any]]:
    if isinstance(conditions, str):
        try:
            parsed = json.loads(conditions)
        except json.JSONDecodeError:
            return []
        conditions = parsed
    if isinstance(conditions, dict):
        conditions = [conditions]
    if not isinstance(conditions, list):
        return []
    return [item for item in conditions if isinstance(item, dict)]


def _is_rank_condition(condition: dict[str, Any]) -> bool:
    op = str(condition.get("op") or condition.get("operator") or "").lower()
    return op in {
        "rank_lt",
        "rank_lte",
        "rank_gt",
        "rank_gte",
        "rank_eq",
        "top",
        "top_n",
        "bottom",
        "bottom_n",
        "排名小于",
        "排名小于等于",
        "排名大于",
        "排名大于等于",
        "排名等于",
        "前",
        "前n",
        "后",
        "后n",
    }


def _condition_rank_limit(condition: dict[str, Any], default: int = 3) -> int:
    for key in ("value", "rank", "n", "max", "limit"):
        value = _to_float(condition.get(key))
        if value is not None:
            return max(1, int(value))
    return default


def _rank_condition_passes(rank: int | None, condition: dict[str, Any]) -> bool:
    if rank is None:
        return False
    op = str(condition.get("op") or condition.get("operator") or "rank_lte").lower()
    limit = _condition_rank_limit(condition)
    if op in {"top", "top_n", "前", "前n", "rank_lte", "排名小于等于"}:
        return rank <= limit
    if op in {"bottom", "bottom_n", "后", "后n"}:
        return rank <= limit
    if op in {"rank_lt", "排名小于"}:
        return rank < limit
    if op in {"rank_gte", "排名大于等于"}:
        return rank >= limit
    if op in {"rank_gt", "排名大于"}:
        return rank > limit
    if op in {"rank_eq", "排名等于"}:
        return rank == limit
    return False


def _parse_listish(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        values = value
    elif isinstance(value, tuple):
        values = list(value)
    else:
        text = _cell_text(value)
        if not text:
            return []
        values = re.split(r"[,，、;；\n]+", text)
    return [_cell_text(item) for item in values if _cell_text(item)]


def _parse_metric_specs(value: Any, default_period: str | None = None) -> list[dict[str, Any]]:
    specs: list[dict[str, Any]] = []
    if value is None:
        return specs
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return specs
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            parsed = [part.strip() for part in re.split(r"[,，、;；\n]+", text) if part.strip()]
    else:
        parsed = value
    if isinstance(parsed, dict):
        parsed = [parsed]
    if not isinstance(parsed, list):
        return specs
    for item in parsed:
        if isinstance(item, dict):
            metric = _cell_text(item.get("metric") or item.get("name") or item.get("col") or item.get("column"))
            reference = item.get("col") or item.get("column") or item.get("reference")
            period = _cell_text(item.get("period") or default_period)
            label = _cell_text(item.get("label") or item.get("name") or metric or reference)
            if metric or reference:
                metric_norm = _normalize_match_text(metric)
                label_norm = _normalize_match_text(label)
                effective_metric = metric
                if metric_norm and label_norm and metric_norm in label_norm and len(label_norm) > len(metric_norm):
                    effective_metric = label
                specs.append(
                    {
                        "metric": metric or None,
                        "effective_metric": effective_metric or metric or None,
                        "reference": reference,
                        "period": period or None,
                        "label": label,
                    }
                )
        else:
            metric = _cell_text(item)
            if metric:
                specs.append({"metric": metric, "effective_metric": metric, "reference": None, "period": default_period, "label": metric})
    return specs


def _locate_metric_specs(
    rows: list[list[Any]],
    *,
    max_col: int,
    header_rows: list[int],
    specs: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    located: list[dict[str, Any]] = []
    for spec in specs:
        match = _locate_column_in_matrix(
            rows,
            max_col=max_col,
            header_rows=header_rows,
            reference=spec.get("reference"),
            metric=spec.get("effective_metric") or spec.get("metric"),
            period=spec.get("period"),
        )
        if match:
            match = {**match, "_header_rows": header_rows, "_max_col": max_col}
        located.append({"spec": spec, "column": match, "status": "found" if match else "column_not_found"})
    return located


def _metric_value_for_row(rows: list[list[Any]], row_number: int, located_item: dict[str, Any]) -> dict[str, Any]:
    column = located_item.get("column")
    spec = located_item.get("spec") or {}
    if not column:
        return {"label": spec.get("label"), "status": "column_not_found", "value": None}
    value = _cell_value(rows, row_number, int(column["index"]))
    descriptor = column.get("descriptor") or " ".join(column.get("header_values") or [])
    normalized = _normalize_metric_number(value, descriptor)
    if normalized is None:
        derived_period = spec.get("period") or _first_period_from_text(
            spec.get("metric"),
            spec.get("label"),
            spec.get("effective_metric"),
            descriptor,
        )
        derived = _derived_ratio_value_for_row(
            rows,
            row_number=row_number,
            metric_text=spec.get("effective_metric") or spec.get("metric") or spec.get("label") or descriptor,
            period=derived_period,
            header_rows=column.get("_header_rows") or [],
            max_col=column.get("_max_col") or len(rows[row_number - 1]) if 1 <= row_number <= len(rows) else 0,
        )
        if derived:
            normalized = derived
    return {
        "label": spec.get("label"),
        "metric": spec.get("metric"),
        "period": spec.get("period"),
        "column": column,
        "value": normalized["value"] if normalized else _to_float(value),
        "display_value": normalized["display_value"] if normalized else _cell_text(value),
        "raw_number": normalized["raw_number"] if normalized else _to_float(value),
        "raw_value": _cell_text(value),
        "normalization": normalized["normalization"] if normalized else "",
    }


def _format_number_for_table(value: Any, *, decimals: int | None = None) -> str:
    number = _to_float(value)
    if number is None:
        return _cell_text(value)
    if decimals is None:
        return f"{number:.6g}"
    places = max(0, min(int(decimals), 8))
    quantum = Decimal("1") if places == 0 else Decimal("1").scaleb(-places)
    rounded = Decimal(str(number)).quantize(quantum, rounding=ROUND_HALF_UP)
    return f"{rounded:.{places}f}"


def _rounded_display_value(value: Any, *, round_decimals: int | None = None, display_decimals: int | None = None) -> str:
    number = _to_float(value)
    if number is None:
        return _cell_text(value)
    if round_decimals is not None:
        number = round(number, max(0, min(int(round_decimals), 8)))
    return _format_number_for_table(number, decimals=display_decimals)


def _markdown_table(headers: list[str], rows: list[list[Any]]) -> str:
    safe_headers = [_cell_text(header).replace("|", "\\|") for header in headers]
    lines = [
        "| " + " | ".join(safe_headers) + " |",
        "| " + " | ".join("---" for _ in safe_headers) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(_cell_text(value).replace("|", "\\|") for value in row) + " |")
    return "\n".join(lines)


def _derived_ratio_value_for_row(
    rows: list[list[Any]],
    *,
    row_number: int,
    metric_text: Any,
    period: str | None,
    header_rows: list[int],
    max_col: int,
) -> dict[str, Any] | None:
    normalized_metric = _normalize_match_text(metric_text)
    if "占" not in normalized_metric or "比" not in normalized_metric:
        return None
    numerator_metric = None
    denominator_metric = None
    if "一年以上" in normalized_metric and "应收总额" in normalized_metric:
        numerator_metric = "一年以上应收"
        denominator_metric = "应收总额"
    if not numerator_metric or not denominator_metric:
        return None
    numerator_col = _locate_column_in_matrix(
        rows,
        max_col=max_col,
        header_rows=header_rows,
        metric=numerator_metric,
        period=period,
    )
    denominator_col = _locate_column_in_matrix(
        rows,
        max_col=max_col,
        header_rows=header_rows,
        metric=denominator_metric,
        period=period,
    )
    if not numerator_col or not denominator_col:
        return None
    numerator = _to_float(_cell_value(rows, row_number, int(numerator_col["index"])))
    denominator = _to_float(_cell_value(rows, row_number, int(denominator_col["index"])))
    if numerator is None or denominator in (None, 0):
        return None
    value = numerator / float(denominator) * 100
    return {
        "value": value,
        "display_value": f"{value:.4g}%",
        "raw_number": value,
        "raw_value": f"{numerator}/{denominator}",
        "normalization": "derived_ratio_percent",
        "derived_from": {
            "numerator_column": numerator_col,
            "denominator_column": denominator_col,
            "numerator": numerator,
            "denominator": denominator,
        },
    }


def _apply_cohort_inference(
    cohort: str | None,
    *,
    cohort_metric: str | None = None,
    cohort_period: str | None = None,
    cohort_min: float | None = None,
) -> tuple[str | None, str | None, float | None]:
    inferred_metric = cohort_metric
    inferred_period = cohort_period
    inferred_min = cohort_min
    cohort_text = _cell_text(cohort)
    if cohort_text and inferred_min is None:
        match = re.search(r"(\d+(?:\.\d+)?)\s*亿", cohort_text)
        if match:
            inferred_min = float(match.group(1))
            if not inferred_metric:
                inferred_metric = "年总收入"
            if not inferred_period:
                year_match = re.search(r"(20\d{2})年?", cohort_text)
                inferred_period = f"{year_match.group(1)}年" if year_match else None
    return inferred_metric, inferred_period, inferred_min


def _domain_cohort_entities(workspace: Path, cohort: str | None) -> list[str]:
    cohort_text = _normalize_match_text(cohort)
    if not cohort_text:
        return []
    knowledge, _knowledge_file = _load_domain_knowledge(workspace)
    if knowledge.get("status") in {"not_found", "invalid_json", "read_error"}:
        return []
    for item in knowledge.get("cohorts") or []:
        names = [item.get("name", ""), *(item.get("aliases") or [])]
        if any(_normalize_match_text(name) and _normalize_match_text(name) in cohort_text for name in names):
            return [_cell_text(entity) for entity in item.get("entities") or [] if _cell_text(entity)]
    return []


@tool_parameters(
    tool_parameters_schema(
        path=StringSchema("Spreadsheet path. Can be absolute, workspace-relative, or a filename under workspace/uploads."),
        value_col=StringSchema("Column reference/name to sort by, such as C, 3, or a header. Optional if metric/period are provided.", nullable=True),
        metric=StringSchema("Metric/header text to sort by if value_col is not explicit.", nullable=True),
        period=StringSchema("Optional period/month header such as 202502 or 2025年2月.", nullable=True),
        sheet=StringSchema("Optional sheet name.", nullable=True),
        k=IntegerSchema(10, description="Number of rows to return.", minimum=1, maximum=100),
        ascending=BooleanSchema(description="Sort ascending. Default false means top/highest first.", default=False),
        entity_col=StringSchema("Optional entity/name column, such as 单位, B, or 2.", nullable=True),
        include_metrics=ArraySchema(
            ObjectSchema(
                metric=StringSchema("Companion metric/header to return from the same ranked row.", nullable=True),
                period=StringSchema("Optional period/month for this companion metric.", nullable=True),
                col=StringSchema("Optional explicit companion column reference such as AA.", nullable=True),
                label=StringSchema("Optional output label.", nullable=True),
            ),
            description="Optional companion metrics to return for each ranked row, e.g. sort by 同比增幅 and include 当期金额.",
            nullable=True,
        ),
        exclude_contains=StringSchema("Comma-separated row text to exclude, for example 合计,市州合计,total.", nullable=True),
        required=["path"],
    )
)
class TableClawTopKTool(Tool):
    """Return top/bottom rows by a numeric spreadsheet column."""

    def __init__(self, workspace: Path | None = None):
        self._workspace = Path(workspace or ".").resolve()

    @classmethod
    def create(cls, ctx: Any) -> Tool:
        return cls(workspace=Path(ctx.workspace))

    @property
    def name(self) -> str:
        return "tableclaw_topk"

    @property
    def description(self) -> str:
        return (
            "Compute top/bottom-k rows from a spreadsheet after locating the numeric metric column. Use for ranking tasks "
            "instead of writing a custom openpyxl sorting script. It normalizes mixed percent encodings such as 0.0902 "
            "and 7.33 in percent-like columns before sorting."
        )

    @property
    def read_only(self) -> bool:
        return True

    async def execute(
        self,
        path: str,
        value_col: str | None = None,
        metric: str | None = None,
        period: str | None = None,
        sheet: str | None = None,
        k: int = 10,
        ascending: bool = False,
        entity_col: str | None = None,
        include_metrics: Any = None,
        exclude_contains: str | None = "合计,市州合计,total",
        **_: Any,
    ) -> str:
        table_path = _resolve_table_path(self._workspace, path)
        if not table_path.exists():
            return f"Error: table file not found: {table_path}"
        schema = _load_or_build_schema(self._workspace, table_path)
        sheet_name = _choose_sheet(schema, sheet)
        selected_sheet, rows, max_row, max_col = _load_sheet_matrix(table_path, sheet_name)
        header_rows = _header_rows_from_matrix(rows, max_col)
        data_start_row = max(header_rows or [1]) + 1
        value_match = _locate_column_in_matrix(
            rows,
            max_col=max_col,
            header_rows=header_rows,
            reference=value_col,
            metric=metric,
            period=period,
        )
        if not value_match:
            return _json_response({"status": "value_column_not_found", "path": str(table_path), "sheet": selected_sheet})
        entity_match = _locate_column_in_matrix(
            rows,
            max_col=max_col,
            header_rows=header_rows,
            reference=entity_col,
            metric="单位" if not entity_col else None,
        )
        entity_col_index = (entity_match or {}).get("index") or 2 if max_col >= 2 else 1
        exclude_terms = _exclude_terms(exclude_contains)
        items: list[dict[str, Any]] = []
        value_col_index = int(value_match["index"])
        value_descriptor = value_match.get("descriptor") or " ".join(value_match.get("header_values") or [])
        companion_specs = _parse_metric_specs(include_metrics, default_period=period)
        located_companions = _locate_metric_specs(
            rows,
            max_col=max_col,
            header_rows=header_rows,
            specs=companion_specs,
        )
        for row_number in range(data_start_row, len(rows) + 1):
            row = rows[row_number - 1]
            row_text = " ".join(_cell_text(value) for value in row if _cell_text(value))
            if not row_text or any(_contains_loose(row_text, term) for term in exclude_terms):
                continue
            value = _cell_value(rows, row_number, value_col_index)
            normalized_number = _normalize_metric_number(value, value_descriptor)
            if normalized_number is None:
                continue
            entity = _cell_text(_cell_value(rows, row_number, int(entity_col_index))) or row_text[:80]
            items.append(
                {
                    "row": row_number,
                    "entity": entity,
                    "value": normalized_number["value"],
                    "display_value": normalized_number["display_value"],
                    "raw_number": normalized_number["raw_number"],
                    "raw_value": _cell_text(value),
                    "normalization": normalized_number["normalization"],
                    "metrics": [_metric_value_for_row(rows, row_number, item) for item in located_companions],
                    "row_text": row_text[:500],
                }
            )
        items.sort(key=lambda item: item["value"], reverse=not _parse_boolish(ascending))
        return _json_response(
            {
                "status": "ok",
                "path": str(table_path),
                "sheet": selected_sheet,
                "header_rows": header_rows,
                "data_start_row": data_start_row,
                "value_column": value_match,
                "include_metrics": located_companions,
                "entity_column": entity_match,
                "ascending": _parse_boolish(ascending),
                "k": k,
                "results": items[: max(1, min(int(k), 100))],
            }
        )


@tool_parameters(
    tool_parameters_schema(
        path=StringSchema("Spreadsheet path. Can be absolute, workspace-relative, or a filename under workspace/uploads."),
        entity=StringSchema("Entity/unit to rank, such as 四川 or 成都."),
        metric=StringSchema("Metric/header text to rank by, such as 应收占收比 or 产数应收占收比."),
        period=StringSchema("Optional period/month header such as 202403 or 2024年3月.", nullable=True),
        sheet=StringSchema("Optional sheet name.", nullable=True),
        ascending=BooleanSchema(
            description="Sort ascending. Leave null to use metric defaults: ratio/risk metrics rank low-to-high unless explicitly requested otherwise.",
            nullable=True,
        ),
        cohort=StringSchema("Optional cohort phrase, such as 200亿省. Numeric 亿 cohorts can be inferred.", nullable=True),
        cohort_metric=StringSchema("Optional column/header used to define the cohort, such as 2023年总收入.", nullable=True),
        cohort_period=StringSchema("Optional period for the cohort column, such as 2023年.", nullable=True),
        cohort_min=NumberSchema(description="Optional minimum cohort value.", nullable=True),
        cohort_max=NumberSchema(description="Optional maximum cohort value.", nullable=True),
        entity_col=StringSchema("Optional entity/name column, such as 单位, B, or 2.", nullable=True),
        exclude_contains=StringSchema("Comma-separated row text to exclude, for example 合计,南方省,北方省,total.", nullable=True),
        required=["path", "entity", "metric"],
    )
)
class TableClawRankTool(Tool):
    """Return an entity's rank for a metric, optionally within a cohort."""

    def __init__(self, workspace: Path | None = None):
        self._workspace = Path(workspace or ".").resolve()

    @classmethod
    def create(cls, ctx: Any) -> Tool:
        return cls(workspace=Path(ctx.workspace))

    @property
    def name(self) -> str:
        return "tableclaw_rank"

    @property
    def description(self) -> str:
        return (
            "Compute an entity's rank for a spreadsheet metric and optional cohort. Use this for questions like "
            "'排名第几', '200亿省排名', or '在某类单位中排名'. It locates columns, normalizes mixed percent encodings "
            "such as 0.0902 vs 7.33, excludes summary rows, and returns overall/cohort rankings. For risk ratio "
            "metrics such as 占收比/占比/比率/率, use ascending=true unless the user explicitly says 从高到低、降序、最高."
        )

    @property
    def read_only(self) -> bool:
        return True

    async def execute(
        self,
        path: str,
        entity: str,
        metric: str,
        period: str | None = None,
        sheet: str | None = None,
        ascending: bool | None = None,
        cohort: str | None = None,
        cohort_metric: str | None = None,
        cohort_period: str | None = None,
        cohort_min: float | None = None,
        cohort_max: float | None = None,
        entity_col: str | None = None,
        exclude_contains: str | None = "合计,南方省,北方省,total",
        **_: Any,
    ) -> str:
        table_path = _resolve_table_path(self._workspace, path)
        if not table_path.exists():
            return f"Error: table file not found: {table_path}"
        schema = _load_or_build_schema(self._workspace, table_path)
        sheet_name = _choose_sheet(schema, sheet)
        selected_sheet, rows, _max_row, max_col = _load_sheet_matrix(table_path, sheet_name)
        header_rows = _header_rows_from_matrix(rows, max_col)
        data_start_row = max(header_rows or [1]) + 1

        value_match = _locate_column_in_matrix(
            rows,
            max_col=max_col,
            header_rows=header_rows,
            metric=metric,
            period=period,
        )
        if not value_match:
            return _json_response({"status": "value_column_not_found", "path": str(table_path), "sheet": selected_sheet, "metric": metric, "period": period})
        rank_match = _locate_rank_column_for_metric(
            rows,
            max_col=max_col,
            header_rows=header_rows,
            metric_column=value_match,
            metric=metric,
            period=period,
        )

        entity_match = _locate_column_in_matrix(
            rows,
            max_col=max_col,
            header_rows=header_rows,
            reference=entity_col,
            metric="单位" if not entity_col else None,
        )
        entity_col_index = (entity_match or {}).get("index") or 2 if max_col >= 2 else 1
        exclude_terms = _exclude_terms(exclude_contains)
        value_col_index = int(value_match["index"])
        value_descriptor = value_match.get("descriptor") or " ".join(value_match.get("header_values") or [])

        inferred_cohort_min = cohort_min
        inferred_cohort_metric = cohort_metric
        inferred_cohort_period = cohort_period
        cohort_text = _cell_text(cohort)
        if cohort_text and inferred_cohort_min is None:
            match = re.search(r"(\d+(?:\.\d+)?)\s*亿", cohort_text)
            if match:
                inferred_cohort_min = float(match.group(1))
                if not inferred_cohort_metric:
                    inferred_cohort_metric = "年总收入"
                if not inferred_cohort_period:
                    year_match = re.search(r"(20\d{2})年?", cohort_text)
                    inferred_cohort_period = f"{year_match.group(1)}年" if year_match else "2023年"

        cohort_match = None
        cohort_col_index = None
        cohort_descriptor = ""
        if inferred_cohort_metric or inferred_cohort_min is not None or cohort_max is not None:
            cohort_match = _locate_column_in_matrix(
                rows,
                max_col=max_col,
                header_rows=header_rows,
                metric=inferred_cohort_metric or "总收入",
                period=inferred_cohort_period,
            )
            if cohort_match:
                cohort_col_index = int(cohort_match["index"])
                cohort_descriptor = cohort_match.get("descriptor") or " ".join(cohort_match.get("header_values") or [])

        records: list[dict[str, Any]] = []
        for row_number in range(data_start_row, len(rows) + 1):
            row = rows[row_number - 1]
            row_text = " ".join(_cell_text(value) for value in row if _cell_text(value))
            if not row_text or any(_contains_loose(row_text, term) for term in exclude_terms):
                continue
            raw_value = _cell_value(rows, row_number, value_col_index)
            normalized = _normalize_metric_number(raw_value, value_descriptor)
            if normalized is None:
                continue
            row_entity = _cell_text(_cell_value(rows, row_number, int(entity_col_index))) or row_text[:80]
            official_rank_raw = _cell_value(rows, row_number, int(rank_match["index"])) if rank_match else None
            official_rank = _parse_rank_value(official_rank_raw)
            cohort_value = None
            cohort_display = ""
            in_cohort = True
            if cohort_col_index:
                raw_cohort = _cell_value(rows, row_number, cohort_col_index)
                cohort_normalized = _normalize_metric_number(raw_cohort, cohort_descriptor)
                cohort_value = cohort_normalized["value"] if cohort_normalized else None
                cohort_display = cohort_normalized["display_value"] if cohort_normalized else _cell_text(raw_cohort)
                if inferred_cohort_min is not None:
                    in_cohort = cohort_value is not None and cohort_value >= float(inferred_cohort_min)
                if cohort_max is not None:
                    in_cohort = in_cohort and cohort_value is not None and cohort_value <= float(cohort_max)
            records.append(
                {
                    "row": row_number,
                    "entity": row_entity,
                    "value": normalized["value"],
                    "display_value": normalized["display_value"],
                    "raw_value": _cell_text(raw_value),
                    "raw_number": normalized["raw_number"],
                    "normalization": normalized["normalization"],
                    "official_rank": official_rank,
                    "official_rank_raw": _cell_text(official_rank_raw),
                    "cohort_value": cohort_value,
                    "cohort_display": cohort_display,
                    "in_cohort": in_cohort,
                }
            )

        effective_ascending = _parse_boolish(ascending) if ascending is not None else (
            True if _is_percent_like_metric(value_descriptor) else False
        )
        reverse = not effective_ascending
        ranked = sorted(records, key=lambda item: item["value"], reverse=reverse)
        for rank, item in enumerate(ranked, start=1):
            item["rank"] = rank
            item["computed_rank"] = rank
            item["recommended_rank"] = item.get("official_rank") or rank
            item["rank_source"] = "official_column" if item.get("official_rank") is not None else "computed"
        cohort_ranked = [item.copy() for item in ranked if item.get("in_cohort")]
        for rank, item in enumerate(cohort_ranked, start=1):
            item["cohort_rank"] = rank
            item["computed_cohort_rank"] = rank
            item["recommended_cohort_rank"] = rank

        def _find_entity(items: list[dict[str, Any]]) -> dict[str, Any] | None:
            for item in items:
                if _contains_loose(item.get("entity"), entity):
                    return item
            return None

        target = _find_entity(ranked)
        cohort_target = _find_entity(cohort_ranked)
        return _json_response(
            {
                "status": "ok" if target else "entity_not_found",
                "path": str(table_path),
                "sheet": selected_sheet,
                "header_rows": header_rows,
                "data_start_row": data_start_row,
                "entity": entity,
                "metric": metric,
                "period": period,
                "ascending": effective_ascending,
                "value_column": value_match,
                "official_rank_column": rank_match,
                "entity_column": entity_match,
                "cohort": {
                    "label": cohort,
                    "metric": inferred_cohort_metric,
                    "period": inferred_cohort_period,
                    "min": inferred_cohort_min,
                    "max": cohort_max,
                    "column": cohort_match,
                    "count": len(cohort_ranked),
                },
                "target": target,
                "cohort_target": cohort_target,
                "ranked": ranked[:50],
                "cohort_ranked": cohort_ranked[:50],
                "next_step": (
                    "Use target.recommended_rank for overall rank when official_rank is present; otherwise use "
                    "target.computed_rank. Use cohort_target.cohort_rank for cohort rank. For percent/ratio risk "
                    "metrics, this tool defaults to low-to-high ranking unless ascending=false was explicitly passed."
                ),
            }
        )


@tool_parameters(
    tool_parameters_schema(
        path=StringSchema("Spreadsheet path. Can be absolute, workspace-relative, or a filename under workspace/uploads."),
        conditions=ArraySchema(
            ObjectSchema(
                col=StringSchema("Explicit column reference/name, optional if metric/period are provided.", nullable=True),
                metric=StringSchema("Metric/header text.", nullable=True),
                period=StringSchema("Optional period/month header.", nullable=True),
                op=StringSchema("Operator: eq, contains, gt, gte, lt, lte, between, ne, rank_lte, rank_gte, top, bottom."),
                value=StringSchema("Comparison value.", nullable=True),
                min=NumberSchema(description="Minimum for between.", nullable=True),
                max=NumberSchema(description="Maximum for between.", nullable=True),
                ascending=BooleanSchema(description="For rank/top conditions only. Default false means top/highest first; set true for low-to-high ranking.", nullable=True),
            ),
            description="List of row conditions; all conditions must pass. Rank conditions such as rank_lte/top can express 前N/排名前三 filters.",
            min_items=1,
            max_items=12,
        ),
        sheet=StringSchema("Optional sheet name.", nullable=True),
        entity_col=StringSchema("Optional entity/name column, such as 单位, B, or 2.", nullable=True),
        select_metrics=ArraySchema(
            ObjectSchema(
                metric=StringSchema("Metric/header to return for each matched row.", nullable=True),
                period=StringSchema("Optional period/month for this output metric.", nullable=True),
                col=StringSchema("Optional explicit output column reference such as AE.", nullable=True),
                label=StringSchema("Optional output label.", nullable=True),
            ),
            description="Optional metrics to return for every matched row, useful for cohort tables and chart datasets.",
            nullable=True,
        ),
        cohort=StringSchema("Optional cohort phrase, such as 200亿省. Uses domain knowledge entities when available, or numeric 亿 inference when configured.", nullable=True),
        cohort_metric=StringSchema("Optional column/header used to define a dynamic cohort, such as 2024年总收入.", nullable=True),
        cohort_period=StringSchema("Optional period for the cohort column, such as 2024年.", nullable=True),
        cohort_min=NumberSchema(description="Optional minimum cohort value.", nullable=True),
        cohort_max=NumberSchema(description="Optional maximum cohort value.", nullable=True),
        exclude_contains=StringSchema("Comma-separated row text to exclude, for example 合计,市州合计,total.", nullable=True),
        limit=IntegerSchema(50, description="Maximum matched rows to return.", minimum=1, maximum=200),
        required=["path", "conditions"],
    )
)
class TableClawFilterTool(Tool):
    """Filter spreadsheet rows with multiple conditions."""

    def __init__(self, workspace: Path | None = None):
        self._workspace = Path(workspace or ".").resolve()

    @classmethod
    def create(cls, ctx: Any) -> Tool:
        return cls(workspace=Path(ctx.workspace))

    @property
    def name(self) -> str:
        return "tableclaw_filter"

    @property
    def description(self) -> str:
        return (
            "Filter spreadsheet rows by multiple conditions, including threshold/range/contains checks. Use for questions "
            "asking which units satisfy criteria or how many rows meet conditions. It also supports rank/top conditions "
            "such as rank_lte/top for 前N/排名前三 filters, optionally within a cohort like 200亿省. For TOP/最高/前N, "
            "default to high-to-low ranking unless the user explicitly asks for lowest/risk-low ordering. Use "
            "select_metrics to return matched rows as a multi-metric dataset for tables/charts."
        )

    @property
    def read_only(self) -> bool:
        return True

    async def execute(
        self,
        path: str,
        conditions: Any,
        sheet: str | None = None,
        entity_col: str | None = None,
        select_metrics: Any = None,
        cohort: str | None = None,
        cohort_metric: str | None = None,
        cohort_period: str | None = None,
        cohort_min: float | None = None,
        cohort_max: float | None = None,
        exclude_contains: str | None = "合计,市州合计,total",
        limit: int = 50,
        **_: Any,
    ) -> str:
        parsed_conditions = _parse_conditions(conditions)
        rank_conditions = [item for item in parsed_conditions if _is_rank_condition(item)]
        value_conditions = [item for item in parsed_conditions if not _is_rank_condition(item)]
        table_path = _resolve_table_path(self._workspace, path)
        if not table_path.exists():
            return f"Error: table file not found: {table_path}"
        schema = _load_or_build_schema(self._workspace, table_path)
        sheet_name = _choose_sheet(schema, sheet)
        selected_sheet, rows, _max_row, max_col = _load_sheet_matrix(table_path, sheet_name)
        header_rows = _header_rows_from_matrix(rows, max_col)
        data_start_row = max(header_rows or [1]) + 1
        located_conditions: list[dict[str, Any]] = []
        for condition in value_conditions:
            match = _locate_column_in_matrix(
                rows,
                max_col=max_col,
                header_rows=header_rows,
                reference=condition.get("col"),
                metric=condition.get("metric"),
                period=condition.get("period"),
            )
            if not match:
                return _json_response({"status": "condition_column_not_found", "condition": condition, "path": str(table_path), "sheet": selected_sheet})
            located_conditions.append({"condition": condition, "column": match})
        located_rank_conditions: list[dict[str, Any]] = []
        for condition in rank_conditions:
            match = _locate_column_in_matrix(
                rows,
                max_col=max_col,
                header_rows=header_rows,
                reference=condition.get("col"),
                metric=condition.get("metric"),
                period=condition.get("period"),
            )
            if not match:
                return _json_response({"status": "rank_condition_column_not_found", "condition": condition, "path": str(table_path), "sheet": selected_sheet})
            located_rank_conditions.append({"condition": condition, "column": match})
        entity_match = _locate_column_in_matrix(
            rows,
            max_col=max_col,
            header_rows=header_rows,
            reference=entity_col,
            metric="单位" if not entity_col else None,
        )
        entity_col_index = (entity_match or {}).get("index") or 2 if max_col >= 2 else 1
        select_specs = _parse_metric_specs(select_metrics)
        located_select_metrics = _locate_metric_specs(
            rows,
            max_col=max_col,
            header_rows=header_rows,
            specs=select_specs,
        )
        domain_cohort_entities = []
        requested_cohort_norms: list[str] = []
        if cohort and not any(value is not None for value in (cohort_metric, cohort_period, cohort_min, cohort_max)):
            domain_cohort_entities = _domain_cohort_entities(self._workspace, cohort)
            requested_cohort_norms = [_normalize_match_text(item) for item in domain_cohort_entities]

        inferred_cohort_metric, inferred_cohort_period, inferred_cohort_min = (None, None, None)
        if not domain_cohort_entities:
            inferred_cohort_metric, inferred_cohort_period, inferred_cohort_min = _apply_cohort_inference(
                cohort,
                cohort_metric=cohort_metric,
                cohort_period=cohort_period,
                cohort_min=cohort_min,
            )
        cohort_match = None
        cohort_col_index = None
        cohort_descriptor = ""
        if inferred_cohort_metric or inferred_cohort_min is not None or cohort_max is not None:
            cohort_match = _locate_column_in_matrix(
                rows,
                max_col=max_col,
                header_rows=header_rows,
                metric=inferred_cohort_metric or "总收入",
                period=inferred_cohort_period,
            )
            if cohort_match:
                cohort_col_index = int(cohort_match["index"])
                cohort_descriptor = cohort_match.get("descriptor") or " ".join(cohort_match.get("header_values") or [])
        exclude_terms = _exclude_terms(exclude_contains)
        candidates: list[dict[str, Any]] = []
        for row_number in range(data_start_row, len(rows) + 1):
            row = rows[row_number - 1]
            row_text = " ".join(_cell_text(value) for value in row if _cell_text(value))
            if not row_text or any(_contains_loose(row_text, term) for term in exclude_terms):
                continue
            entity = _cell_text(_cell_value(rows, row_number, int(entity_col_index))) or row_text[:80]
            entity_norm = _normalize_match_text(entity)
            if entity_norm in {"单位", "名称", "省份", "市州", "区县"}:
                continue
            in_cohort = True
            cohort_value = None
            cohort_display = ""
            if requested_cohort_norms:
                in_cohort = any(norm and norm in entity_norm for norm in requested_cohort_norms)
            if cohort_col_index:
                raw_cohort = _cell_value(rows, row_number, cohort_col_index)
                cohort_normalized = _normalize_metric_number(raw_cohort, cohort_descriptor)
                cohort_value = cohort_normalized["value"] if cohort_normalized else None
                cohort_display = cohort_normalized["display_value"] if cohort_normalized else _cell_text(raw_cohort)
                if inferred_cohort_min is not None:
                    in_cohort = cohort_value is not None and cohort_value >= float(inferred_cohort_min)
                if cohort_max is not None:
                    in_cohort = in_cohort and cohort_value is not None and cohort_value <= float(cohort_max)
            if not in_cohort:
                continue
            candidates.append(
                {
                    "row": row_number,
                    "entity": entity,
                    "cohort_value": cohort_value,
                    "cohort_display": cohort_display,
                    "row_text": row_text[:500],
                }
            )

        rank_maps: list[dict[int, int]] = []
        for item in located_rank_conditions:
            col_index = int(item["column"]["index"])
            descriptor = item["column"].get("descriptor") or " ".join(item["column"].get("header_values") or [])
            rank_items = []
            for candidate in candidates:
                raw_value = _cell_value(rows, int(candidate["row"]), col_index)
                normalized = _normalize_metric_number(raw_value, descriptor)
                if normalized is None:
                    continue
                rank_items.append({"row": int(candidate["row"]), "value": normalized["value"]})
            ascending = item["condition"].get("ascending")
            if ascending is None:
                op = str(item["condition"].get("op") or item["condition"].get("operator") or "").lower()
                ascending = op in {"bottom", "bottom_n", "后", "后n"}
            rank_items.sort(key=lambda value: value["value"], reverse=not _parse_boolish(ascending))
            rank_maps.append({int(value["row"]): rank for rank, value in enumerate(rank_items, start=1)})

        matches: list[dict[str, Any]] = []
        for candidate in candidates:
            row_number = int(candidate["row"])
            checks = []
            passed = True
            for item in located_conditions:
                col_index = int(item["column"]["index"])
                actual = _cell_value(rows, row_number, col_index)
                descriptor = item["column"].get("descriptor") or " ".join(item["column"].get("header_values") or [])
                normalized_actual = _normalize_metric_number(actual, descriptor)
                ok = _condition_passes(actual, item["condition"], descriptor)
                checks.append(
                    {
                        "column": item["column"],
                        "op": item["condition"].get("op") or "eq",
                        "expected": item["condition"].get("value"),
                        "actual": _cell_text(actual),
                        "actual_normalized": normalized_actual["value"] if normalized_actual else None,
                        "actual_display": normalized_actual["display_value"] if normalized_actual else _cell_text(actual),
                        "normalization": normalized_actual["normalization"] if normalized_actual else "",
                        "passed": ok,
                    }
                )
                if not ok:
                    passed = False
                    break
            if passed:
                for index, item in enumerate(located_rank_conditions):
                    row_rank = rank_maps[index].get(row_number) if index < len(rank_maps) else None
                    ok = _rank_condition_passes(row_rank, item["condition"])
                    checks.append(
                        {
                            "column": item["column"],
                            "op": item["condition"].get("op") or "rank_lte",
                            "expected": item["condition"].get("value") or item["condition"].get("rank") or item["condition"].get("n"),
                            "rank": row_rank,
                            "passed": ok,
                        }
                    )
                    if not ok:
                        passed = False
                        break
            if not passed:
                continue
            matches.append(
                {
                    "row": row_number,
                    "entity": candidate["entity"],
                    "cohort_value": candidate.get("cohort_value"),
                    "cohort_display": candidate.get("cohort_display"),
                    "checks": checks,
                    "metrics": [_metric_value_for_row(rows, row_number, item) for item in located_select_metrics],
                    "row_text": candidate["row_text"],
                }
            )
        return _json_response(
            {
                "status": "ok",
                "path": str(table_path),
                "sheet": selected_sheet,
                "header_rows": header_rows,
                "data_start_row": data_start_row,
                "conditions": located_conditions,
                "rank_conditions": located_rank_conditions,
                "entity_column": entity_match,
                "select_metrics": located_select_metrics,
                "cohort": {
                    "label": cohort,
                    "domain_entities": domain_cohort_entities,
                    "metric": inferred_cohort_metric,
                    "period": inferred_cohort_period,
                    "min": inferred_cohort_min,
                    "max": cohort_max,
                    "column": cohort_match,
                    "candidate_count": len(candidates),
                },
                "matched_count": len(matches),
                "results": matches[: max(1, min(int(limit), 200))],
            }
        )


@tool_parameters(
    tool_parameters_schema(
        path=StringSchema("Spreadsheet path. Can be absolute, workspace-relative, or a filename under workspace/uploads."),
        metrics=ArraySchema(
            ObjectSchema(
                metric=StringSchema("Metric/header to return, such as 应收总额 or 占收比.", nullable=True),
                period=StringSchema("Optional period/month for this metric, such as 202509.", nullable=True),
                col=StringSchema("Optional explicit output column reference such as M.", nullable=True),
                label=StringSchema("Optional output row label.", nullable=True),
            ),
            description="Metrics to extract for each entity. Output is row x metric and chart-ready transposed table.",
            min_items=1,
            max_items=12,
        ),
        entities=StringSchema(
            "Optional comma-separated entity/unit names. If omitted, all data rows are used after filters/exclusions.",
            nullable=True,
        ),
        conditions=ArraySchema(
            ObjectSchema(
                col=StringSchema("Explicit condition column reference/name, optional if metric/period are provided.", nullable=True),
                metric=StringSchema("Condition metric/header text.", nullable=True),
                period=StringSchema("Optional period/month for condition column.", nullable=True),
                op=StringSchema("Operator: eq, contains, gt, gte, lt, lte, between, ne."),
                value=StringSchema("Comparison value.", nullable=True),
                min=NumberSchema(description="Minimum for between.", nullable=True),
                max=NumberSchema(description="Maximum for between.", nullable=True),
            ),
            description="Optional filters; all conditions must pass before output.",
            nullable=True,
            max_items=12,
        ),
        sheet=StringSchema("Optional sheet name.", nullable=True),
        entity_col=StringSchema("Optional entity/name column, such as 单位, B, or 2.", nullable=True),
        sort_by=StringSchema("Optional metric label/name or entity to sort rows. Use first metric when omitted and ascending/descending is requested.", nullable=True),
        ascending=BooleanSchema(description="Sort rows ascending by sort_by. Default null preserves source/requested order.", nullable=True),
        transpose=BooleanSchema(description="Include chart-ready table with metric rows and entity columns.", default=True),
        cohort=StringSchema("Optional cohort phrase, such as 200亿省. Numeric 亿 cohorts are inferred dynamically from an income column.", nullable=True),
        cohort_metric=StringSchema("Optional column/header used to define the cohort, such as 2024年总收入.", nullable=True),
        cohort_period=StringSchema("Optional period for the cohort column, such as 2024年.", nullable=True),
        cohort_min=NumberSchema(description="Optional minimum cohort value.", nullable=True),
        cohort_max=NumberSchema(description="Optional maximum cohort value.", nullable=True),
        exclude_contains=StringSchema("Comma-separated row text to exclude, for example 合计,市州合计,total.", nullable=True),
        limit=IntegerSchema(100, description="Maximum rows/entities to return.", minimum=1, maximum=300),
        display_decimals=IntegerSchema(
            description="Decimals for chart_table values. Defaults to 1 for chart-ready tables unless explicitly set.",
            nullable=True,
            minimum=0,
            maximum=8,
        ),
        required=["path", "metrics"],
    )
)
class TableClawExtractMatrixTool(Tool):
    """Extract a clean entity-by-metric matrix for chart/table answers."""

    def __init__(self, workspace: Path | None = None):
        self._workspace = Path(workspace or ".").resolve()

    @classmethod
    def create(cls, ctx: Any) -> Tool:
        return cls(workspace=Path(ctx.workspace))

    @property
    def name(self) -> str:
        return "tableclaw_extract_matrix"

    @property
    def description(self) -> str:
        return (
            "Extract chart-ready table data from one spreadsheet as entities x metrics. Use for chart/table questions, "
            "especially when the user asks to draw/compare several provinces/cities or multiple indicators. It locates "
            "multi-row headers, supports explicit entity lists, dynamic numeric cohorts, filters, sorting, and percent "
            "normalization, then returns both row-oriented records and a transposed markdown table. Include qualifiers "
            "such as 基础/产数/一年以上 in metric or label; for 大省/主要大省 without an explicit list, use cohort='200亿省' "
            "or another dynamic income cohort instead of hand-picking entities. If the user gives a single month, return "
            "that month's snapshot table even if the requested output is a chart/trend chart; only use cross-period tools "
            "when the user explicitly asks for multiple months or a time range."
        )

    @property
    def read_only(self) -> bool:
        return True

    async def execute(
        self,
        path: str,
        metrics: Any,
        entities: str | None = None,
        conditions: Any = None,
        sheet: str | None = None,
        entity_col: str | None = None,
        sort_by: str | None = None,
        ascending: bool | None = None,
        transpose: bool = True,
        cohort: str | None = None,
        cohort_metric: str | None = None,
        cohort_period: str | None = None,
        cohort_min: float | None = None,
        cohort_max: float | None = None,
        exclude_contains: str | None = "合计,南方省,北方省,市州合计,total",
        limit: int = 100,
        display_decimals: int | None = None,
        **_: Any,
    ) -> str:
        if display_decimals is None:
            display_decimals = 1
        table_path = _resolve_table_path(self._workspace, path)
        if not table_path.exists():
            return f"Error: table file not found: {table_path}"
        schema = _load_or_build_schema(self._workspace, table_path)
        sheet_name = _choose_sheet(schema, sheet)
        selected_sheet, rows, _max_row, max_col = _load_sheet_matrix(table_path, sheet_name)
        header_rows = _header_rows_from_matrix(rows, max_col)
        data_start_row = max(header_rows or [1]) + 1

        metric_specs = _parse_metric_specs(metrics)
        if not metric_specs:
            return _json_response({"status": "no_metrics", "path": str(table_path), "sheet": selected_sheet})
        located_metrics = _locate_metric_specs(
            rows,
            max_col=max_col,
            header_rows=header_rows,
            specs=metric_specs,
        )
        missing_metrics = [item for item in located_metrics if not item.get("column")]
        if missing_metrics:
            return _json_response(
                {
                    "status": "metric_column_not_found",
                    "path": str(table_path),
                    "sheet": selected_sheet,
                    "missing_metrics": missing_metrics,
                    "located_metrics": located_metrics,
                }
            )

        entity_match = _locate_column_in_matrix(
            rows,
            max_col=max_col,
            header_rows=header_rows,
            reference=entity_col,
            metric="单位" if not entity_col else None,
        )
        entity_col_index = (entity_match or {}).get("index") or 2 if max_col >= 2 else 1

        parsed_conditions = _parse_conditions(conditions)
        located_conditions: list[dict[str, Any]] = []
        for condition in parsed_conditions:
            match = _locate_column_in_matrix(
                rows,
                max_col=max_col,
                header_rows=header_rows,
                reference=condition.get("col"),
                metric=condition.get("metric"),
                period=condition.get("period"),
            )
            if not match:
                return _json_response({"status": "condition_column_not_found", "condition": condition, "path": str(table_path), "sheet": selected_sheet})
            located_conditions.append({"condition": condition, "column": match})

        requested_entities = _parse_listish(entities)
        requested_entity_norms = [_normalize_match_text(item) for item in requested_entities]
        domain_cohort_entities = []
        if not requested_entities and cohort and not any(
            value is not None for value in (cohort_metric, cohort_period, cohort_min, cohort_max)
        ):
            domain_cohort_entities = _domain_cohort_entities(self._workspace, cohort)
            if domain_cohort_entities:
                requested_entities = domain_cohort_entities
                requested_entity_norms = [_normalize_match_text(item) for item in requested_entities]

        inferred_cohort_metric, inferred_cohort_period, inferred_cohort_min = (None, None, None)
        if not domain_cohort_entities:
            inferred_cohort_metric, inferred_cohort_period, inferred_cohort_min = _apply_cohort_inference(
                cohort,
                cohort_metric=cohort_metric,
                cohort_period=cohort_period,
                cohort_min=cohort_min,
            )
        cohort_match = None
        cohort_col_index = None
        cohort_descriptor = ""
        if inferred_cohort_metric or inferred_cohort_min is not None or cohort_max is not None:
            cohort_match = _locate_column_in_matrix(
                rows,
                max_col=max_col,
                header_rows=header_rows,
                metric=inferred_cohort_metric or "总收入",
                period=inferred_cohort_period,
            )
            if cohort_match:
                cohort_col_index = int(cohort_match["index"])
                cohort_descriptor = cohort_match.get("descriptor") or " ".join(cohort_match.get("header_values") or [])

        exclude_terms = _exclude_terms(exclude_contains)
        records: list[dict[str, Any]] = []
        seen_requested: set[str] = set()
        for row_number in range(data_start_row, len(rows) + 1):
            row = rows[row_number - 1]
            row_text = " ".join(_cell_text(value) for value in row if _cell_text(value))
            if not row_text or any(_contains_loose(row_text, term) for term in exclude_terms):
                continue
            entity_name = _cell_text(_cell_value(rows, row_number, int(entity_col_index))) or row_text[:80]
            if not _cell_text(_cell_value(rows, row_number, int(entity_col_index))):
                continue
            if _normalize_match_text(entity_name) in {"单位", "名称", "省份", "市州", "区县"}:
                continue
            if requested_entities:
                entity_norm = _normalize_match_text(entity_name)
                matched_index = next((idx for idx, norm in enumerate(requested_entity_norms) if norm and norm in entity_norm), None)
                if matched_index is None:
                    continue
                requested_key = requested_entities[matched_index]
                if requested_key in seen_requested:
                    continue
                seen_requested.add(requested_key)
            else:
                matched_index = None

            condition_checks = []
            passed = True
            for item in located_conditions:
                col_index = int(item["column"]["index"])
                actual = _cell_value(rows, row_number, col_index)
                descriptor = item["column"].get("descriptor") or " ".join(item["column"].get("header_values") or [])
                ok = _condition_passes(actual, item["condition"], descriptor)
                condition_checks.append(
                    {
                        "column": item["column"],
                        "op": item["condition"].get("op") or "eq",
                        "expected": item["condition"].get("value"),
                        "actual": _cell_text(actual),
                        "passed": ok,
                    }
                )
                if not ok:
                    passed = False
                    break
            if not passed:
                continue

            cohort_value = None
            cohort_display = ""
            in_cohort = True
            if cohort_col_index:
                raw_cohort = _cell_value(rows, row_number, cohort_col_index)
                cohort_normalized = _normalize_metric_number(raw_cohort, cohort_descriptor)
                cohort_value = cohort_normalized["value"] if cohort_normalized else None
                cohort_display = cohort_normalized["display_value"] if cohort_normalized else _cell_text(raw_cohort)
                if inferred_cohort_min is not None:
                    in_cohort = cohort_value is not None and cohort_value >= float(inferred_cohort_min)
                if cohort_max is not None:
                    in_cohort = in_cohort and cohort_value is not None and cohort_value <= float(cohort_max)
            if not in_cohort:
                continue

            values = [_metric_value_for_row(rows, row_number, item) for item in located_metrics]
            record = {
                "row": row_number,
                "source_order": len(records) + 1,
                "requested_order": matched_index + 1 if matched_index is not None else None,
                "entity": requested_entities[matched_index] if requested_entities and matched_index is not None else entity_name,
                "sheet_entity": entity_name,
                "cohort_value": cohort_value,
                "cohort_display": cohort_display,
                "condition_checks": condition_checks,
                "metrics": values,
                "row_text": row_text[:500],
            }
            records.append(record)

        sort_label = _cell_text(sort_by)
        if ascending is None and not sort_label and not requested_entities and not cohort:
            ascending = False
        if ascending is not None or sort_label:
            reverse = not _parse_boolish(ascending)

            def _sort_value(record: dict[str, Any]) -> Any:
                if sort_label and _contains_loose(record.get("entity"), sort_label):
                    return record.get("entity") or ""
                for metric_item in record.get("metrics") or []:
                    if not sort_label or _contains_loose(metric_item.get("label"), sort_label) or _contains_loose(metric_item.get("metric"), sort_label):
                        value = metric_item.get("value")
                        return value if value is not None else float("-inf")
                return float("-inf")

            records.sort(key=_sort_value, reverse=reverse)

        records = records[: max(1, min(int(limit), 300))]
        metric_labels = [
            _cell_text((item.get("spec") or {}).get("label"))
            or _cell_text((item.get("spec") or {}).get("metric"))
            or _cell_text((item.get("column") or {}).get("descriptor"))
            for item in located_metrics
        ]
        row_table: list[list[str]] = []
        for record in records:
            row_table.append(
                [record["entity"]]
                + [_format_number_for_table((metric_item or {}).get("value"), decimals=display_decimals) for metric_item in record.get("metrics") or []]
            )
        transposed_table: list[list[str]] = []
        if transpose:
            for metric_index, label in enumerate(metric_labels):
                transposed_table.append(
                    [label]
                    + [
                        _format_number_for_table((record.get("metrics") or [])[metric_index].get("value"), decimals=display_decimals)
                        for record in records
                    ]
                )
        chart_markdown = _markdown_table(["单位", *[record["entity"] for record in records]], transposed_table) if transpose else ""
        row_markdown = _markdown_table(["单位", *metric_labels], row_table)
        answer_markdown = chart_markdown or row_markdown
        return _json_response(
            {
                "status": "ok",
                "path": str(table_path),
                "sheet": selected_sheet,
                "header_rows": header_rows,
                "data_start_row": data_start_row,
                "entity_column": entity_match,
                "metrics": located_metrics,
                "conditions": located_conditions,
                "cohort": {
                    "label": cohort,
                    "domain_entities": domain_cohort_entities,
                    "metric": inferred_cohort_metric,
                    "period": inferred_cohort_period,
                    "min": inferred_cohort_min,
                    "max": cohort_max,
                    "column": cohort_match,
                },
                "requested_entities": requested_entities,
                "matched_count": len(records),
                "answer_markdown": answer_markdown,
                "answer_instruction": (
                    "For chart-generation eval answers, copy answer_markdown/chart_table.markdown directly. "
                    "Do not rebuild a different orientation or change decimals unless the user explicitly asked for that format."
                ),
                "records": records,
                "row_table": {
                    "headers": ["单位", *metric_labels],
                    "rows": row_table,
                    "markdown": row_markdown,
                },
                "chart_table": {
                    "headers": ["单位", *[record["entity"] for record in records]],
                    "rows": transposed_table,
                    "markdown": chart_markdown,
                },
                "next_step": (
                    "For chart-generation eval answers, copy answer_markdown/chart_table.markdown directly when the output "
                    "is for plotting; copy row_table.markdown only when the user explicitly asks for entity rows. If this "
                    "table contains the requested period/entities/metrics, answer directly with it and do not continue "
                    "exploring or add extra trend/time-series tables."
                ),
            }
        )


@tool_parameters(
    tool_parameters_schema(
        path=StringSchema("Spreadsheet path. Can be absolute, workspace-relative, or a filename under workspace/uploads."),
        entity=StringSchema("Entity/unit/province/city name to extract, for example 四川 or 成都.", nullable=True),
        metric=StringSchema("Metric/header text to extract across periods.", nullable=True),
        periods=StringSchema("Comma-separated periods or range text, for example 202501,202502 or 2025年1-12月.", nullable=True),
        period_start=StringSchema("Optional start period such as 202501.", nullable=True),
        period_end=StringSchema("Optional end period such as 202512.", nullable=True),
        sheet=StringSchema("Optional sheet name.", nullable=True),
        entity_col=StringSchema("Optional entity/name column, such as 单位, B, or 2.", nullable=True),
        exclude_contains=StringSchema("Comma-separated row text to exclude, for example 合计,市州合计,total.", nullable=True),
        required=["path"],
    )
)
class TableClawExtractSeriesTool(Tool):
    """Extract period series values for an entity and metric."""

    def __init__(self, workspace: Path | None = None):
        self._workspace = Path(workspace or ".").resolve()

    @classmethod
    def create(cls, ctx: Any) -> Tool:
        return cls(workspace=Path(ctx.workspace))

    @property
    def name(self) -> str:
        return "tableclaw_extract_series"

    @property
    def description(self) -> str:
        return (
            "Extract a month/period series for an entity and metric from a spreadsheet with multi-row headers. Use for trend "
            "tables, cross-month comparisons, or extracting the same metric for several comma-separated entities instead "
            "of manually scanning each column."
        )

    @property
    def read_only(self) -> bool:
        return True

    async def execute(
        self,
        path: str,
        entity: str | None = None,
        metric: str | None = None,
        periods: str | None = None,
        period_start: str | None = None,
        period_end: str | None = None,
        sheet: str | None = None,
        entity_col: str | None = None,
        exclude_contains: str | None = "合计,市州合计,total",
        **_: Any,
    ) -> str:
        table_path = _resolve_table_path(self._workspace, path)
        if not table_path.exists():
            return f"Error: table file not found: {table_path}"
        schema = _load_or_build_schema(self._workspace, table_path)
        sheet_name = _choose_sheet(schema, sheet)
        selected_sheet, rows, _max_row, max_col = _load_sheet_matrix(table_path, sheet_name)
        header_rows = _header_rows_from_matrix(rows, max_col)
        data_start_row = max(header_rows or [1]) + 1
        parsed_periods = _parse_periods(periods, period_start, period_end)
        if not parsed_periods and periods:
            parsed_periods = [part.strip() for part in re.split(r"[,，、\s]+", periods) if part.strip()]
        requested_entities = _parse_listish(entity)
        row_matches: list[dict[str, Any]] = []
        if requested_entities:
            for requested_entity in requested_entities:
                for row_match in _locate_row(
                    rows,
                    data_start_row=data_start_row,
                    entity=requested_entity,
                    entity_col=entity_col,
                    exclude_contains=exclude_contains,
                ):
                    row_matches.append({**row_match, "entity": requested_entity})
        else:
            row_matches = _locate_row(
                rows,
                data_start_row=data_start_row,
                entity=entity,
                entity_col=entity_col,
                exclude_contains=exclude_contains,
            )
        if entity and not row_matches:
            return _json_response({"status": "entity_not_found", "entity": entity, "path": str(table_path), "sheet": selected_sheet})
        target_rows = row_matches[:100] if entity else _locate_row(
            rows,
            data_start_row=data_start_row,
            entity=None,
            entity_col=entity_col,
            exclude_contains=exclude_contains,
        )[:10]
        series: list[dict[str, Any]] = []
        periods_to_scan = parsed_periods or [None]
        located_columns: list[dict[str, Any]] = []
        for period in periods_to_scan:
            column = _locate_column_in_matrix(
                rows,
                max_col=max_col,
                header_rows=header_rows,
                metric=metric,
                period=period,
            )
            if not column:
                series.append({"period": period, "status": "column_not_found"})
                continue
            located_columns.append(column)
            for row_match in target_rows:
                value = _cell_value(rows, int(row_match["row"]), int(column["index"]))
                descriptor = column.get("descriptor") or " ".join(column.get("header_values") or [])
                normalized = _normalize_metric_number(value, descriptor)
                series.append(
                    {
                        "period": period,
                        "row": row_match["row"],
                        "entity": row_match.get("entity") or entity or row_match["row_text"][:80],
                        "column": column,
                        "value": normalized["value"] if normalized else _to_float(value),
                        "display_value": normalized["display_value"] if normalized else _cell_text(value),
                        "raw_number": normalized["raw_number"] if normalized else _to_float(value),
                        "raw_value": _cell_text(value),
                        "normalization": normalized["normalization"] if normalized else "",
                    }
                )
        return _json_response(
            {
                "status": "ok",
                "path": str(table_path),
                "sheet": selected_sheet,
                "header_rows": header_rows,
                "data_start_row": data_start_row,
                "entity": entity,
                "metric": metric,
                "periods": parsed_periods,
                "target_rows": target_rows,
                "located_columns": located_columns,
                "series": series,
            }
        )


@tool_parameters(
    tool_parameters_schema(
        metric=StringSchema("Metric/header text to extract across monthly files, such as 应收总额 or 产数应收总额."),
        entity=StringSchema("Entity/unit/province/city name to extract, for example 四川 or 全省.", nullable=True),
        file_pattern=StringSchema(
            "Optional uploaded filename pattern containing {period}, e.g. 全国各省份数据-通报应收总额_{period}.xlsx.",
            nullable=True,
        ),
        table_family=StringSchema(
            "Optional filename keywords used when file_pattern is omitted, e.g. 通报应收总额 or 市州应收账款情况表.",
            nullable=True,
        ),
        periods=StringSchema("Comma-separated periods or range text, for example 202501,202502 or 2025年1-12月.", nullable=True),
        period_start=StringSchema("Optional start period such as 202501.", nullable=True),
        period_end=StringSchema("Optional end period such as 202512.", nullable=True),
        sheet=StringSchema("Optional sheet name.", nullable=True),
        entity_col=StringSchema("Optional entity/name column, such as 单位, B, or 2.", nullable=True),
        exclude_contains=StringSchema("Comma-separated row text to exclude, for example 合计,市州合计,total.", nullable=True),
        compute=StringSchema(
            "Optional derived calculation: none, mom_percent, difference, percent_change. mom_percent compares each period to previous month.",
            enum=["none", "mom_percent", "difference", "percent_change"],
            nullable=True,
        ),
        display_decimals=IntegerSchema(
            description="Optional decimals for derived/table values. Example: 2 for 环比增幅 tables.",
            nullable=True,
            minimum=0,
            maximum=8,
        ),
        required=["metric"],
    )
)
class TableClawTimeSeriesTool(Tool):
    """Extract one metric/entity across monthly spreadsheet files."""

    def __init__(self, workspace: Path | None = None):
        self._workspace = Path(workspace or ".").resolve()

    @classmethod
    def create(cls, ctx: Any) -> Tool:
        return cls(workspace=Path(ctx.workspace))

    @property
    def name(self) -> str:
        return "tableclaw_time_series"

    @property
    def description(self) -> str:
        return (
            "Extract the same entity and metric across many monthly uploaded spreadsheets. Use for 1-12月趋势、跨期比较、"
            "环比/增长率 tables instead of manually opening each workbook. It resolves period filenames, locates the same "
            "metric column per file, normalizes percentages, and can compute MoM percent changes. Do not use it for a "
            "single-month snapshot chart unless the user explicitly asks for multiple periods. For 应收账款绝对值/应收总额 "
            "trend questions, use 通报应收总额 tables with metric='应收总额'; do not substitute 长账龄/应收账款余额 unless "
            "the user explicitly asks for 余额、账龄 or 一年以上."
        )

    @property
    def read_only(self) -> bool:
        return True

    async def execute(
        self,
        metric: str,
        entity: str | None = None,
        file_pattern: str | None = None,
        table_family: str | None = None,
        periods: str | None = None,
        period_start: str | None = None,
        period_end: str | None = None,
        sheet: str | None = None,
        entity_col: str | None = None,
        exclude_contains: str | None = "合计,南方省,北方省,市州合计,total",
        compute: str | None = "none",
        display_decimals: int | None = None,
        **_: Any,
    ) -> str:
        metric_norm = _normalize_match_text(metric)
        pattern_norm = _normalize_match_text(file_pattern)
        family_norm = _normalize_match_text(table_family)
        if (
            ("长账龄" in pattern_norm or "长账龄" in family_norm)
            and "余额" in metric_norm
            and any(term in metric_norm for term in ("绝对值", "应收总额", "应收账款总额"))
        ):
            return _json_response(
                {
                    "status": "metric_family_mismatch",
                    "metric": metric,
                    "file_pattern": file_pattern,
                    "table_family": table_family,
                    "reason": "The requested metric looks like receivable total/absolute value, but the chosen table family is aging balance.",
                    "recommended_next_step": (
                        "Use 全国各省份数据-通报应收总额_{period}.xlsx with metric='应收总额' for 应收账款绝对值/应收总额; "
                        "use 长账龄 only for 一年以上、账龄、月末余额 or explicitly requested 应收账款余额."
                    ),
                }
            )
        display_round_decimals = 1 if compute in {"mom_percent", "percent_change"} else None
        if display_decimals is None and compute in {"mom_percent", "percent_change"}:
            display_decimals = 2
        parsed_periods = _parse_periods(periods, period_start, period_end)
        if not parsed_periods and periods:
            parsed_periods = [part.strip() for part in re.split(r"[,，、\s]+", periods) if part.strip()]
        if not parsed_periods:
            return _json_response({"status": "no_periods", "metric": metric, "entity": entity})

        points: list[dict[str, Any]] = []
        needed_periods = list(parsed_periods)
        if compute in {"mom_percent", "difference", "percent_change"}:
            previous = _previous_period(parsed_periods[0])
            if previous and previous not in needed_periods:
                needed_periods = [previous, *needed_periods]

        for period in needed_periods:
            table_path = _resolve_period_table(
                self._workspace,
                file_pattern=file_pattern,
                table_family=table_family,
                period=period,
            )
            if not table_path:
                points.append({"period": period, "status": "table_not_found"})
                continue
            schema = _load_or_build_schema(self._workspace, table_path)
            sheet_name = _choose_sheet(schema, sheet)
            selected_sheet, rows, _max_row, max_col = _load_sheet_matrix(table_path, sheet_name)
            header_rows = _header_rows_from_matrix(rows, max_col)
            data_start_row = max(header_rows or [1]) + 1
            column = _locate_column_in_matrix(
                rows,
                max_col=max_col,
                header_rows=header_rows,
                metric=metric,
                period=period,
            ) or _locate_column_in_matrix(
                rows,
                max_col=max_col,
                header_rows=header_rows,
                metric=metric,
            )
            if not column:
                points.append({"period": period, "path": str(table_path), "sheet": selected_sheet, "status": "column_not_found"})
                continue
            row_matches = _locate_row(
                rows,
                data_start_row=data_start_row,
                entity=entity,
                entity_col=entity_col,
                exclude_contains=exclude_contains,
            )
            if entity and not row_matches:
                points.append({"period": period, "path": str(table_path), "sheet": selected_sheet, "status": "entity_not_found"})
                continue
            target_row = row_matches[0] if row_matches else None
            if not target_row:
                points.append({"period": period, "path": str(table_path), "sheet": selected_sheet, "status": "row_not_found"})
                continue
            raw_value = _cell_value(rows, int(target_row["row"]), int(column["index"]))
            descriptor = column.get("descriptor") or " ".join(column.get("header_values") or [])
            normalized = _normalize_metric_number(raw_value, descriptor)
            points.append(
                {
                    "period": period,
                    "period_label": _period_label(period),
                    "status": "ok",
                    "path": str(table_path),
                    "filename": table_path.name,
                    "sheet": selected_sheet,
                    "row": target_row["row"],
                    "entity": entity or target_row["row_text"][:80],
                    "column": column,
                    "value": normalized["value"] if normalized else _to_float(raw_value),
                    "display_value": normalized["display_value"] if normalized else _cell_text(raw_value),
                    "raw_number": normalized["raw_number"] if normalized else _to_float(raw_value),
                    "raw_value": _cell_text(raw_value),
                    "normalization": normalized["normalization"] if normalized else "",
                }
            )

        values_by_period = {point["period"]: point for point in points if point.get("status") == "ok"}
        derived: list[dict[str, Any]] = []
        for period in parsed_periods:
            point = values_by_period.get(period)
            value = point.get("value") if point else None
            derived_value = value
            base_period = None
            base_value = None
            if compute in {"mom_percent", "difference", "percent_change"}:
                base_period = _previous_period(period)
                base_point = values_by_period.get(base_period or "")
                base_value = base_point.get("value") if base_point else None
                if value is not None and base_value not in (None, 0):
                    if compute in {"mom_percent", "percent_change"}:
                        derived_value = (float(value) - float(base_value)) / float(base_value) * 100
                    elif compute == "difference":
                        derived_value = float(value) - float(base_value)
                else:
                    derived_value = None
            derived.append(
                {
                    "period": period,
                    "period_label": _period_label(period),
                    "value": value,
                    "display_value": _rounded_display_value(value, display_decimals=display_decimals),
                    "base_period": base_period,
                    "base_value": base_value,
                    "derived_value": derived_value,
                    "derived_display": _rounded_display_value(
                        derived_value,
                        round_decimals=display_round_decimals,
                        display_decimals=display_decimals,
                    ),
                    "source": point,
                    "status": "ok" if point else "missing",
                }
            )

        if compute in {"mom_percent", "percent_change"}:
            row_label = f"{metric}环比增幅（%）"
            value_key = "derived_display"
        elif compute == "difference":
            row_label = f"{metric}差值"
            value_key = "derived_display"
        else:
            row_label = metric
            value_key = "display_value"
        headers = ["单位", *[_period_label(period) for period in parsed_periods]]
        table_rows = [[row_label, *[(item.get(value_key) or "") for item in derived]]]
        chart_table = {
            "headers": headers,
            "rows": table_rows,
            "markdown": _markdown_table(headers, table_rows),
        }
        return _json_response(
            {
                "status": "ok",
                "metric": metric,
                "entity": entity,
                "periods": parsed_periods,
                "file_pattern": file_pattern,
                "table_family": table_family,
                "compute": compute or "none",
                "answer_markdown": chart_table["markdown"],
                "chart_table": chart_table,
                "answer_instruction": (
                    "Copy answer_markdown/chart_table.markdown directly for the final answer. "
                    "Do not recompute from points or add raw absolute-value columns unless the user explicitly asked for them."
                ),
                "points": points,
                "derived": derived,
                "used_files": [point.get("filename") for point in points if point.get("filename")],
                "next_step": "Use answer_markdown/chart_table.markdown exactly for trend-table/chart eval answers; do not recompute from raw points.",
            }
        )


@tool_parameters(
    tool_parameters_schema(
        query=StringSchema("Full user question, including dates, metrics, units, and scope."),
        top_k=IntegerSchema(8, description="Maximum uploaded table candidates to return.", minimum=1, maximum=20),
        rebuild_index=BooleanSchema(description="Rebuild the workspace table index from uploads before retrieval.", default=False),
        required=["query"],
    )
)
class TableClawRetrieveTablesTool(Tool):
    """Retrieve uploaded spreadsheet candidates from the TableClaw workspace."""

    def __init__(self, workspace: Path | None = None):
        self._workspace = Path(workspace or ".").resolve()

    @classmethod
    def create(cls, ctx: Any) -> Tool:
        return cls(workspace=Path(ctx.workspace))

    @property
    def name(self) -> str:
        return "tableclaw_retrieve_tables"

    @property
    def description(self) -> str:
        return (
            "Retrieve likely relevant uploaded spreadsheet files from workspace/uploads. "
            "Use this first when the user asks about tables/spreadsheets but does not provide an exact file path, "
            "or when multiple uploaded tables may be relevant. It returns candidate file paths plus scope, subject, "
            "month, sheet/schema previews, scores, and match reasons; it does not read full spreadsheet contents. "
            "After retrieval, call tableclaw_inspect on the most relevant candidate before writing spreadsheet code."
        )

    @property
    def read_only(self) -> bool:
        return True

    async def execute(self, query: str, top_k: int = 8, rebuild_index: bool = False, **_: Any) -> str:
        upload_dir = self._workspace / "uploads"
        index_file = self._workspace / "table_index" / "tables.jsonl"
        if rebuild_index or not index_file.exists():
            index = _build_index(upload_dir, index_file, self._workspace)
        else:
            index = _load_index(index_file) or _build_index(upload_dir, index_file, self._workspace)

        index = _enrich_index_with_catalog(self._workspace, index)
        intent = _parse_retrieval_intent(query)
        candidates = _retrieve(query, index, top_k=top_k)
        table_groups = _retrieve_groups(query, index, top_k=5)
        compact_candidates = []
        for rank, item in enumerate(candidates, start=1):
            compact_candidates.append(
                {
                    "rank": rank,
                    "table_id": item["table_id"],
                    "filename": item["filename"],
                    "path": item["path"],
                    "score": item["score"],
                    "scope": item.get("scope"),
                    "subject": item.get("subject"),
                    "month": item.get("month"),
                    "keywords": item.get("keywords"),
                    "reasons": item.get("reasons"),
                    "risks": item.get("risks"),
                    "fit": item.get("fit"),
                    "description_status": item.get("description_status"),
                    "short_description": item.get("short_description"),
                    "row_grain": item.get("row_grain"),
                    "important_metrics": (item.get("important_metrics") or [])[:12],
                    "can_answer": (item.get("can_answer") or [])[:6],
                    "description_path": item.get("description_path"),
                    "profile_path": item.get("profile_path"),
                    "clean_view_path": item.get("clean_view_path"),
                    "sheets": [
                        {
                            "sheet": sheet.get("sheet"),
                            "max_row": sheet.get("max_row"),
                            "max_column": sheet.get("max_column"),
                            "header_candidates": sheet.get("header_candidates", [])[:2],
                            "preview_text": sheet.get("preview_text", "")[:360],
                        }
                        for sheet in item.get("sheets", [])[:2]
                    ],
                }
            )
        return json.dumps(
            {
                "query": query,
                "workspace": str(self._workspace),
                "upload_dir": str(upload_dir),
                "index_file": str(index_file),
                "catalog_file": str(_catalog_file(self._workspace)),
                "catalog_available": bool(_load_catalog(self._workspace)),
                "indexed_tables": len(index),
                "retrieval_version": "v5-structured-intent",
                "intent": intent,
                "top_k": top_k,
                "candidates": compact_candidates,
                "table_groups": table_groups,
                "next_step": (
                    "Use intent/fit/risks to choose candidate paths. For multi-month or trend tasks, prefer table_groups. "
                    "For industrial finance shorthand such as 200亿省、欠费、小微ICT、一年以上、市州排名、营业收现率、预收 or 保证金, call tableclaw_domain_knowledge for cohort/indicator/table-family guidance before final extraction. "
                    "For 应收账款绝对值/应收总额/应收账款情况, prefer 通报应收总额 tables; use 长账龄 tables only when the query mentions 一年以上、账龄、月末余额 or 应收账款余额. "
                    "For rank/ranking questions such as 排名第几 or 200亿省排名, prefer tableclaw_rank/tableclaw_topk before custom code. "
                    "For TopK questions that also ask for corresponding amounts, call tableclaw_topk with include_metrics. "
                    "For chart/table questions that need a clean bottom table, call tableclaw_extract_matrix with metrics, entities/cohort, sorting, and display_decimals; include metric qualifiers like 基础/产数/一年以上. "
                    "When only one month is specified, chart/trend-chart wording still usually needs one snapshot bottom table; use time_series only for explicit multi-month ranges. "
                    "When the user says 大省/主要大省 and does not list entities, prefer a dynamic income cohort such as cohort='200亿省' if the table has 年总收入. "
                    "For multi-month trend or MoM/环比 questions across monthly files, call tableclaw_time_series with file_pattern/table_family and compute. "
                    "Catalog descriptions are planning context; validate exact cells with tableclaw_inspect/locate/extract/rank/topk/filter/matrix/time_series."
                ),
            },
            ensure_ascii=False,
            indent=2,
        )
