#!/usr/bin/env python3
"""Run a TableClaw retrieval + skill orchestration smoke test.

This runner simulates a user-uploaded table workspace:

1. copy source tables into workspace/uploads;
2. build a lightweight filename + workbook-preview index;
3. retrieve top-k candidate tables from a user question;
4. pass only retrieved candidates to Nanobot and record skill/tool traces.

It is intentionally a smoke harness, not a final accuracy benchmark.
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import os
import re
import shutil
import time
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from nanobot.nanobot import Nanobot
from run_eval import CONFIGS, ROOT, _usage, extract_tool_timeline


SOURCE_TABLE_DIR = ROOT / "test_table"
WORKSPACE_DIR = ROOT / "workspace"
UPLOAD_DIR = WORKSPACE_DIR / "uploads"
INDEX_DIR = WORKSPACE_DIR / "table_index"
INDEX_FILE = INDEX_DIR / "tables.jsonl"
MANIFEST_FILE = UPLOAD_DIR / "upload_manifest.jsonl"
TASK_FILE = ROOT / "eval_test/test_dataset/raw_eval_cleaned.jsonl"
DEFAULT_JSON_OUTPUT = ROOT / "eval_test/results/retrieval_smoke/latest_retrieval_smoke.json"
DEFAULT_MD_OUTPUT = ROOT / "docs/实验评测/retrieval-smoke.md"

TABLE_EXTENSIONS = {".xlsx", ".xls", ".csv", ".tsv"}
QUESTION_TERMS = [
    "欠费",
    "总欠费",
    "未列收",
    "已列收",
    "一年以上",
    "小微ICT",
    "ICT",
    "应收",
    "应收账款",
    "应收占收比",
    "预收",
    "占收比",
    "同比",
    "增幅",
    "营业收现率",
    "营业现金比率",
    "长账龄",
    "保证金",
    "公有池",
    "私有池",
    "大额长账",
    "市州",
    "区县",
    "全省",
    "200亿省",
    "四川",
    "成都",
    "绵阳",
    "自贡",
    "达州",
    "乐山",
    "巴中",
]


def compact(value: Any, max_len: int = 220) -> str:
    text = json.dumps(value, ensure_ascii=False) if not isinstance(value, str) else value
    text = " ".join(text.split()).replace("|", "\\|")
    return text if len(text) <= max_len else text[: max_len - 3] + "..."


def stable_id(path: Path) -> str:
    return "tbl_" + hashlib.sha1(path.name.encode("utf-8")).hexdigest()[:10]


def iter_source_tables(source_dir: Path) -> list[Path]:
    files = []
    for path in source_dir.iterdir():
        if not path.is_file():
            continue
        if path.name.startswith("~$") or path.name.startswith("."):
            continue
        if path.suffix.lower() in TABLE_EXTENSIONS:
            files.append(path)
    return sorted(files, key=lambda item: item.name)


def prepare_uploads(source_dir: Path = SOURCE_TABLE_DIR, upload_dir: Path = UPLOAD_DIR) -> list[dict[str, Any]]:
    upload_dir.mkdir(parents=True, exist_ok=True)
    manifest: list[dict[str, Any]] = []
    for source in iter_source_tables(source_dir):
        target = upload_dir / source.name
        if not target.exists() or target.stat().st_size != source.stat().st_size:
            shutil.copy2(source, target)
        manifest.append(
            {
                "table_id": stable_id(target),
                "filename": target.name,
                "uploaded_path": str(target),
                "source_path": str(source),
                "size_bytes": target.stat().st_size,
                "simulated_upload": True,
            }
        )
    MANIFEST_FILE.write_text(
        "\n".join(json.dumps(item, ensure_ascii=False) for item in manifest) + "\n",
        encoding="utf-8",
    )
    return manifest


def infer_scope(filename: str) -> str:
    if "区县" in filename:
        return "county"
    if "市州" in filename:
        return "city"
    if "全国各省份" in filename or "200亿省" in filename:
        return "province"
    return "unknown"


def infer_subject(filename: str) -> str:
    candidates = [
        "欠费",
        "通报应收总额",
        "应收账款",
        "营业收现率",
        "营业现金比率",
        "长账龄",
        "保证金",
        "公有池",
        "私有池",
        "大额长账",
    ]
    return ",".join(term for term in candidates if term in filename) or "unknown"


def extract_filename_month(filename: str) -> str | None:
    match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", filename)
    return match.group(0) if match else None


def read_workbook_preview(path: Path, *, max_sheets: int = 2, max_rows: int = 8, max_cols: int = 12) -> list[dict[str, Any]]:
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return []
    previews: list[dict[str, Any]] = []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - smoke metadata should be best effort
        return [{"error": str(exc)}]
    try:
        for sheet_name in workbook.sheetnames[:max_sheets]:
            sheet = workbook[sheet_name]
            rows: list[list[str]] = []
            for row in sheet.iter_rows(
                min_row=1,
                max_row=min(max_rows, sheet.max_row or max_rows),
                max_col=min(max_cols, sheet.max_column or max_cols),
                values_only=True,
            ):
                values = [str(value).strip() for value in row if value not in (None, "")]
                if values:
                    rows.append(values)
            preview_text = " ".join(" ".join(row) for row in rows)
            previews.append(
                {
                    "sheet": sheet_name,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "preview_text": preview_text[:1600],
                }
            )
    finally:
        workbook.close()
    return previews


def build_index(upload_dir: Path = UPLOAD_DIR, index_file: Path = INDEX_FILE) -> list[dict[str, Any]]:
    index_file.parent.mkdir(parents=True, exist_ok=True)
    records: list[dict[str, Any]] = []
    for path in iter_source_tables(upload_dir):
        previews = read_workbook_preview(path)
        preview_text = " ".join(item.get("preview_text", "") for item in previews)
        records.append(
            {
                "table_id": stable_id(path),
                "filename": path.name,
                "path": str(path),
                "suffix": path.suffix.lower(),
                "size_bytes": path.stat().st_size,
                "mtime": int(path.stat().st_mtime),
                "scope": infer_scope(path.name),
                "subject": infer_subject(path.name),
                "month": extract_filename_month(path.name),
                "sheets": previews,
                "keywords": sorted(set(term for term in QUESTION_TERMS if term in path.name or term in preview_text)),
            }
        )
    index_file.write_text(
        "\n".join(json.dumps(record, ensure_ascii=False) for record in records) + "\n",
        encoding="utf-8",
    )
    return records


def load_index(index_file: Path = INDEX_FILE) -> list[dict[str, Any]]:
    if not index_file.exists():
        return []
    return [json.loads(line) for line in index_file.read_text(encoding="utf-8").splitlines() if line.strip()]


def load_tasks(task_file: Path = TASK_FILE) -> list[dict[str, Any]]:
    return [json.loads(line) for line in task_file.read_text(encoding="utf-8").splitlines() if line.strip()]


def extract_question_months(question: str) -> list[str]:
    months: set[str] = set()
    for year, start, end in re.findall(r"(20\d{2})年\s*(\d{1,2})\s*[-至到]\s*(\d{1,2})月", question):
        for month in range(int(start), int(end) + 1):
            if 1 <= month <= 12:
                months.add(f"{year}{month:02d}")
    for year, month in re.findall(r"(20\d{2})年\s*(\d{1,2})月", question):
        months.add(f"{year}{int(month):02d}")
    for raw in re.findall(r"20\d{2}(?:0[1-9]|1[0-2])", question):
        months.add(raw)
    return sorted(months)


def extract_question_terms(question: str) -> list[str]:
    terms = [term for term in QUESTION_TERMS if term.lower() in question.lower()]
    for token in re.findall(r"[A-Za-z0-9]+", question):
        if len(token) >= 2:
            terms.append(token)
    return sorted(set(terms), key=lambda item: (-len(item), item))


def score_record(question: str, record: dict[str, Any]) -> tuple[float, list[str]]:
    filename = record["filename"]
    subject = record.get("subject") or ""
    scope = record.get("scope") or ""
    keywords = set(record.get("keywords") or [])
    preview_text = " ".join(item.get("preview_text", "") for item in record.get("sheets") or [])
    haystack = f"{filename} {subject} {scope} {' '.join(keywords)} {preview_text}"
    terms = extract_question_terms(question)
    months = extract_question_months(question)

    score = 0.0
    reasons: list[str] = []
    for term in terms:
        if term in filename:
            score += 8
            reasons.append(f"filename:{term}")
        elif term in subject:
            score += 5
            reasons.append(f"subject:{term}")
        elif term in keywords:
            score += 4
            reasons.append(f"keyword:{term}")
        elif term in preview_text:
            score += 2
            reasons.append(f"preview:{term}")

    if months:
        if record.get("month") in months:
            score += 10
            reasons.append(f"month:{record['month']}")
        elif any(month[:4] in filename for month in months):
            score += 2
            reasons.append("year-match")
        elif not record.get("month") and any(term in haystack for term in ("欠费", "应收", "长账龄")):
            score += 8
            reasons.append("year-series-ledger")
    if "欠费" in question and "欠费" in haystack:
        score += 18
        reasons.append("domain:arrears")
    if "小微ICT" in question and ("小微ICT" in haystack or "小微" in haystack or "ICT" in haystack):
        score += 8
        reasons.append("domain:ict")
    if ("应收" in question or "预收" in question or "占收比" in question) and any(
        term in haystack for term in ("应收", "预收", "占收比", "通报应收总额")
    ):
        score += 12
        reasons.append("domain:receivable")
    if "一年以上" in question and any(term in haystack for term in ("一年以上", "长账龄", "欠费")):
        score += 10
        reasons.append("domain:aging")
    if "全省" in question or "200亿省" in question:
        if scope == "province":
            score += 6
            reasons.append("scope:province")
        if "全国各省份" in filename:
            score += 4
            reasons.append("filename:province")
    if "市州" in question and scope == "city":
        score += 6
        reasons.append("scope:city")
    if "区县" in question and scope == "county":
        score += 6
        reasons.append("scope:county")
    if "画" in question or "图" in question:
        if any(term in haystack for term in ("欠费", "应收", "营业", "长账龄")):
            score += 1
            reasons.append("chart-compatible")
    return score, reasons[:10]


def retrieve(question: str, index: list[dict[str, Any]], *, top_k: int) -> list[dict[str, Any]]:
    scored = []
    for record in index:
        score, reasons = score_record(question, record)
        if score > 0:
            scored.append({**record, "score": round(score, 2), "reasons": reasons})
    scored.sort(key=lambda item: (-item["score"], item["filename"]))
    return scored[:top_k]


def select_tasks(tasks: list[dict[str, Any]], *, limit: int, task_ids: list[str] | None) -> list[dict[str, Any]]:
    if task_ids:
        wanted = set(task_ids)
        return [task for task in tasks if task["id"] in wanted]

    buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for task in tasks:
        if task.get("data_eval_ready"):
            buckets[task.get("task_type") or "unknown"].append(task)

    selected: list[dict[str, Any]] = []
    quotas = [("chart_generation", 4), ("table_qa", 3), ("ranking_qa", 3)]
    for task_type, quota in quotas:
        selected.extend(buckets.get(task_type, [])[:quota])
    if len(selected) < limit:
        already = {task["id"] for task in selected}
        selected.extend(task for task in tasks if task["id"] not in already and task.get("data_eval_ready"))
    return selected[:limit]


def render_prompt(task: dict[str, Any], retrieved: list[dict[str, Any]]) -> str:
    candidates = []
    for idx, item in enumerate(retrieved, start=1):
        candidates.append(
            f"{idx}. {item['path']}\n"
            f"   table_id={item['table_id']} score={item['score']} "
            f"scope={item.get('scope')} subject={item.get('subject')} "
            f"month={item.get('month') or '-'} reasons={', '.join(item.get('reasons') or [])}"
        )
    candidate_text = "\n".join(candidates) if candidates else "未召回到候选表。"
    visual_note = (
        "这是画图/可视化类任务。本轮 smoke eval 只要求输出可用于绘图的底层数据表，"
        "不需要真正生成图片文件。"
        if task.get("requires_visual_artifact")
        else "请直接回答问题，并尽量说明使用了哪些候选表。"
    )
    return f"""你是 TableClaw 的表格 workflow agent。用户已经上传了一批表格，系统先做了表格召回。

用户问题：
{task['question']}

TableClaw retrieval 召回的候选表如下（这是系统检索结果，不是标准答案）：
{candidate_text}

执行要求：
1. 先判断任务是否需要读取表格相关 skill；如涉及 xlsx/csv 读取、结构识别、清洗、验证、图表底层数据，请按需读取合适的 SKILL.md。
2. 优先使用上面的候选表；如果候选表不足，请明确说明缺口。
3. {visual_note}
4. 输出要简洁，最后列出“使用的候选表文件名”和“是否成功完成”。"""


async def run_one(task: dict[str, Any], retrieved: list[dict[str, Any]], *, mode: str) -> dict[str, Any]:
    bot = Nanobot.from_config(CONFIGS[mode])
    prompt = render_prompt(task, retrieved)
    started = time.time()
    try:
        result = await bot.run(
            prompt,
            session_key=f"sdk:retrieval-smoke-{task['id']}-{mode}-{int(started)}",
        )
        answer = result.content
        messages = result.messages
        tools_used = result.tools_used
        error = None
    except Exception as exc:  # pragma: no cover - smoke harness records failure per case
        answer = ""
        messages = []
        tools_used = []
        error = repr(exc)
    finally:
        usage = dict(getattr(bot._loop, "_last_usage", {}) or {})
        await bot._loop.close_mcp()

    elapsed_ms = int((time.time() - started) * 1000)
    timeline = extract_tool_timeline(messages)
    skill_events = [event for event in timeline if event.get("is_tracked_skill_read")]
    selected_skills: list[str] = []
    for event in skill_events:
        skill = event.get("skill_read")
        if skill and skill not in selected_skills:
            selected_skills.append(skill)

    return {
        "task_id": task["id"],
        "task_type": task.get("task_type"),
        "question": task["question"],
        "mode": mode,
        "elapsed_ms": elapsed_ms,
        "usage": usage,
        "retrieved": [
            {
                "rank": idx,
                "table_id": item["table_id"],
                "filename": item["filename"],
                "path": item["path"],
                "score": item["score"],
                "scope": item.get("scope"),
                "subject": item.get("subject"),
                "month": item.get("month"),
                "reasons": item.get("reasons"),
            }
            for idx, item in enumerate(retrieved, start=1)
        ],
        "tools_used": tools_used,
        "tool_timeline": timeline,
        "skill_selected": bool(skill_events),
        "selected_skills": selected_skills,
        "skill_read_sequence": [event.get("skill_read") for event in skill_events],
        "answer_preview": answer[:1200],
        "error": error,
    }


def build_summary(payload: dict[str, Any]) -> str:
    results = payload["results"]
    skill_count = sum(1 for item in results if item.get("skill_selected"))
    total_tokens = sum(_usage(item, "total_tokens") for item in results)
    avg_tokens = round(total_tokens / len(results)) if results else 0
    avg_elapsed = round(sum(item["elapsed_ms"] for item in results) / len(results) / 1000, 1) if results else 0
    skills_seen = sorted({skill for item in results for skill in item.get("selected_skills") or []})
    lines = [
        "# TableClaw Retrieval Smoke",
        "",
        f"> Generated at: {payload['generated_at']}",
        "",
        "## Scope",
        "",
        "本轮用于跑通 `用户上传表格 -> table index -> question retrieval -> Nanobot skill workflow -> trace/usage log` 的编排链路。",
        "当前不作为最终准确率结论；候选表来自检索，不把 gold table path 显式写进 prompt。",
        "",
        "## Key Observations",
        "",
        f"- {payload['indexed_tables']} 张工业表已模拟上传到 `workspace/uploads/`，并生成 `workspace/table_index/tables.jsonl` 文件级索引。",
        f"- {skill_count}/{len(results)} case 触发了 table skill read；触发过 `{', '.join(skills_seen) or '-'}`。",
        f"- 总 token 约 {total_tokens:,}，平均每题约 {avg_tokens:,}；平均耗时约 {avg_elapsed} 秒。",
        "- 编排链路已经跑通：先召回候选表，再由 Nanobot 读取 skill、选择候选表、执行 openpyxl 脚本、输出答案。",
        "- 当前瓶颈很清楚：模型仍在反复摸表头和列位置。下一步应优先做 schema cache、sheet/column retrieval、`tableclaw_locate_column`、`tableclaw_extract_series` 和 `tableclaw_topk` 等工具。",
        "",
        "## Workspace Uploads",
        "",
        f"- Upload dir: `{payload['upload_dir']}`",
        f"- Indexed tables: `{payload['indexed_tables']}`",
        f"- Top-k per task: `{payload['top_k']}`",
        "",
        "## Results",
        "",
        "| Task | Type | Retrieved top1 | Skills read | Skill sequence | Total tokens | Prompt | Completion | Tools | Elapsed ms | Error |",
        "| --- | --- | --- | --- | --- | ---: | ---: | ---: | --- | ---: | --- |",
    ]
    for item in results:
        top1 = item["retrieved"][0]["filename"] if item.get("retrieved") else "-"
        skills = ",".join(item.get("selected_skills") or []) or "-"
        sequence = " -> ".join(item.get("skill_read_sequence") or []) or "-"
        tools = ",".join(item.get("tools_used") or []) or "-"
        lines.append(
            "| "
            + " | ".join(
                [
                    item["task_id"],
                    item.get("task_type") or "-",
                    top1,
                    f"`{skills}`",
                    f"`{sequence}`",
                    str(_usage(item, "total_tokens")),
                    str(_usage(item, "prompt_tokens")),
                    str(_usage(item, "completion_tokens")),
                    tools,
                    str(item["elapsed_ms"]),
                    compact(item.get("error") or "-"),
                ]
            )
            + " |"
        )

    lines.extend(["", "## Retrieval Details", ""])
    for item in results:
        lines.extend(
            [
                f"### {item['task_id']}",
                "",
                f"Question: {item['question']}",
                "",
                "| Rank | Table | Score | Scope | Subject | Month | Reasons |",
                "| ---: | --- | ---: | --- | --- | --- | --- |",
            ]
        )
        for candidate in item.get("retrieved") or []:
            lines.append(
                "| "
                + " | ".join(
                    [
                        str(candidate["rank"]),
                        candidate["filename"],
                        str(candidate["score"]),
                        str(candidate.get("scope") or "-"),
                        str(candidate.get("subject") or "-"),
                        str(candidate.get("month") or "-"),
                        compact(", ".join(candidate.get("reasons") or [])),
                    ]
                )
                + " |"
            )
        lines.extend(["", "Answer preview:", "", "```text", item["answer_preview"], "```", ""])
    return "\n".join(lines)


async def async_main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=10)
    parser.add_argument("--top-k", type=int, default=8)
    parser.add_argument("--mode", choices=CONFIGS.keys(), default="skill-on")
    parser.add_argument("--task-id", action="append")
    parser.add_argument("--skip-upload", action="store_true")
    parser.add_argument("--skip-index", action="store_true")
    parser.add_argument("--dry-run", action="store_true", help="Only prepare uploads, index, and show retrieval results.")
    parser.add_argument("--json-output", default=str(DEFAULT_JSON_OUTPUT.relative_to(ROOT)))
    parser.add_argument("--md-output", default=str(DEFAULT_MD_OUTPUT.relative_to(ROOT)))
    args = parser.parse_args()

    if not args.skip_upload:
        manifest = prepare_uploads()
    else:
        manifest = [
            json.loads(line)
            for line in MANIFEST_FILE.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ] if MANIFEST_FILE.exists() else []

    if not args.skip_index:
        index = build_index()
    else:
        index = load_index()

    tasks = select_tasks(load_tasks(), limit=args.limit, task_ids=args.task_id)
    retrievals = [(task, retrieve(task["question"], index, top_k=args.top_k)) for task in tasks]

    if args.dry_run:
        for task, candidates in retrievals:
            print(f"\n{task['id']}\t{task.get('task_type')}\t{task['question']}")
            for idx, candidate in enumerate(candidates[: args.top_k], start=1):
                print(f"  {idx}. {candidate['filename']}\t{candidate['score']}\t{candidate.get('reasons')}")
        return

    if not os.environ.get("DASHSCOPE_API_KEY"):
        os.environ["DASHSCOPE_API_KEY"] = "${DASHSCOPE_API_KEY}"

    results: list[dict[str, Any]] = []
    for task, candidates in retrievals:
        print(f"Running {task['id']} / {args.mode} with {len(candidates)} retrieved tables...", flush=True)
        results.append(await run_one(task, candidates, mode=args.mode))

    payload = {
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        "mode": args.mode,
        "top_k": args.top_k,
        "upload_dir": str(UPLOAD_DIR),
        "manifest_file": str(MANIFEST_FILE),
        "index_file": str(INDEX_FILE),
        "uploaded_tables": len(manifest),
        "indexed_tables": len(index),
        "results": results,
    }

    json_path = ROOT / args.json_output
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    md_path = ROOT / args.md_output
    md_path.parent.mkdir(parents=True, exist_ok=True)
    summary = build_summary(payload)
    md_path.write_text(summary, encoding="utf-8")
    print(summary)
    print(f"\nJSON: {json_path}")
    print(f"Markdown: {md_path}")


if __name__ == "__main__":
    asyncio.run(async_main())
