#!/usr/bin/env python3
"""Build a sampled query-variant eval set from the synonym expansion workbook."""

from __future__ import annotations

import argparse
import hashlib
import json
import random
from pathlib import Path

from openpyxl import load_workbook

from import_gold_cases import classify, normalize_text, parse_markdown_table


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "原始数据_图表_同义词随机替换扩展5倍.xlsx"
DEFAULT_OUTPUT = ROOT / "eval_test/test_dataset/query_variants_100.jsonl"


def stable_id(index: int, question: str, answer: str) -> str:
    digest = hashlib.sha1(f"{question}\n---\n{answer}".encode("utf-8")).hexdigest()[:10]
    return f"query_variant_{index:03d}_{digest}"


def build_record(index: int, question: str, answer: str, *, sheet: str, source_row: int, block_index: int) -> dict:
    info = classify(question, answer)
    answer_table = parse_markdown_table(answer)
    return {
        "id": stable_id(index, question, answer),
        "gold_case_index": index,
        "source": {
            "raw_file": DEFAULT_INPUT.name,
            "sheet": sheet,
            "row": source_row,
            "block_index": block_index,
        },
        "question": question,
        "ground_truth": answer,
        "gold_answer": answer,
        "ground_truth_table": answer_table,
        "recommended_eval_mode": "chart_data_table_only" if info["requires_visual_artifact"] else "structured_table_qa",
        "data_eval_ready": True,
        "visual_eval_ready": False,
        "retrieval_eval_ready": False,
        "table_id": None,
        "table_path": None,
        **info,
        "evaluation": {
            "type": "query_variant_gold_answer_reference",
            "method": "same_as_gold_cases_llm_judge",
            "notes": "Sampled from synonym-expanded query workbook. Prompt uses question only; gold_answer is evaluator-only.",
        },
    }


def import_cases(
    input_path: Path,
    output_path: Path,
    *,
    sheet_name: str,
    source_rows: int,
    group_size: int,
    sample_size: int,
    seed: int,
) -> list[dict]:
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found. Available: {workbook.sheetnames}")
    sheet = workbook[sheet_name]
    headers = [normalize_text(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    if headers[:2] != ["问题", "标准答案"]:
        raise ValueError(f"Expected first two headers to be 问题/标准答案, got {headers[:2]}")

    rows: list[tuple[int, str, str]] = []
    for source_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        question = normalize_text(values[0] if len(values) > 0 else "")
        answer = normalize_text(values[1] if len(values) > 1 else "")
        if not question and not answer:
            continue
        if not question or not answer:
            raise ValueError(f"Row {source_row} has missing question or answer")
        rows.append((source_row, question, answer))
        if len(rows) >= source_rows:
            break

    if len(rows) < group_size:
        raise ValueError(f"Need at least {group_size} rows, got {len(rows)}")

    rng = random.Random(seed)
    blocks = [rows[idx : idx + group_size] for idx in range(0, len(rows), group_size)]
    blocks = [block for block in blocks if len(block) == group_size]
    if sample_size > len(blocks):
        raise ValueError(f"Requested {sample_size} samples, but only {len(blocks)} full blocks are available")

    records = []
    for index, block in enumerate(blocks[:sample_size], start=1):
        source_row, question, answer = rng.choice(block)
        records.append(
            build_record(
                index,
                question,
                answer,
                sheet=sheet_name,
                source_row=source_row,
                block_index=index,
            )
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        "\n".join(json.dumps(record, ensure_ascii=False) for record in records) + "\n",
        encoding="utf-8",
    )
    return records


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--sheet", default="图表")
    parser.add_argument("--source-rows", type=int, default=500)
    parser.add_argument("--group-size", type=int, default=5)
    parser.add_argument("--sample-size", type=int, default=100)
    parser.add_argument("--seed", type=int, default=20260615)
    args = parser.parse_args()

    input_path = args.input if args.input.is_absolute() else ROOT / args.input
    output_path = args.output if args.output.is_absolute() else ROOT / args.output
    records = import_cases(
        input_path,
        output_path,
        sheet_name=args.sheet,
        source_rows=args.source_rows,
        group_size=args.group_size,
        sample_size=args.sample_size,
        seed=args.seed,
    )
    counts: dict[str, int] = {}
    for record in records:
        counts[record["task_type"]] = counts.get(record["task_type"], 0) + 1
    print(
        json.dumps(
            {
                "output": str(output_path),
                "records": len(records),
                "sheet": args.sheet,
                "source_rows": args.source_rows,
                "group_size": args.group_size,
                "sample_size": args.sample_size,
                "seed": args.seed,
                "task_type": counts,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
