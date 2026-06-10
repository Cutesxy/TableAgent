#!/usr/bin/env python3
"""Import curated gold spreadsheet QA/chart cases from an xlsx file."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "eval_test/test_dataset/source/测试case抽样.xlsx"
DEFAULT_OUTPUT = ROOT / "eval_test/test_dataset/gold_cases.jsonl"


TASK_PATTERNS = {
    "chart": re.compile(r"图|图表|绘制|画|柱状|折线|双轴|组合图|可视化|横轴|纵轴"),
    "ranking": re.compile(r"最高|最低|排名|前\\d+|后\\d+|排序|最大|最小|从高到低|从低到高"),
    "trend": re.compile(r"1-12月|逐月|时间序列|月度|趋势|环比"),
    "filter": re.compile(r"哪些|有没有|同时满足|超过|低于|为负|下滑|筛选"),
}


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\ufeff", "").strip()
    text = re.sub(r"\r\n?", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    return text


def stable_id(index: int, question: str, answer: str) -> str:
    digest = hashlib.sha1(f"{question}\n---\n{answer}".encode("utf-8")).hexdigest()[:10]
    return f"gold_case_{index:03d}_{digest}"


def parse_markdown_table(markdown: str) -> list[dict[str, str]]:
    lines = [line.strip() for line in markdown.splitlines() if line.strip()]
    table_lines = [line for line in lines if line.startswith("|") and line.endswith("|")]
    if len(table_lines) < 2:
        return []

    def split_row(line: str) -> list[str]:
        return [cell.strip() for cell in line.strip("|").split("|")]

    header = split_row(table_lines[0])
    rows: list[dict[str, str]] = []
    for line in table_lines[1:]:
        cells = split_row(line)
        if cells and all(re.fullmatch(r":?-{2,}:?", cell or "") for cell in cells):
            continue
        rows.append({name or f"col_{i+1}": cells[i] if i < len(cells) else "" for i, name in enumerate(header)})
    return rows


def classify(question: str, answer: str) -> dict[str, Any]:
    facets = [name for name, pattern in TASK_PATTERNS.items() if pattern.search(question)]
    if "chart" in facets:
        task_type = "chart_generation"
    elif "trend" in facets:
        task_type = "trend_table"
    elif "ranking" in facets:
        task_type = "ranking_qa"
    elif "filter" in facets:
        task_type = "filter_qa"
    else:
        task_type = "table_qa"
    return {
        "task_type": task_type,
        "facets": facets,
        "requires_visual_artifact": "chart" in facets,
        "requires_file_edit": "chart" in facets,
        "ground_truth_format": "markdown_table" if parse_markdown_table(answer) else "text",
    }


def build_record(index: int, question: str, answer: str) -> dict[str, Any]:
    info = classify(question, answer)
    answer_table = parse_markdown_table(answer)
    visual_note = "chart_data_table_only" if info["requires_visual_artifact"] else "structured_table_qa"
    return {
        "id": stable_id(index, question, answer),
        "gold_case_index": index,
        "source": {
            "raw_file": "测试case抽样.xlsx",
            "sheet": "Sheet1",
            "row": index + 1,
        },
        "question": question,
        "ground_truth": answer,
        "gold_answer": answer,
        "ground_truth_table": answer_table,
        "recommended_eval_mode": visual_note,
        "data_eval_ready": True,
        "visual_eval_ready": False,
        "retrieval_eval_ready": False,
        "table_id": None,
        "table_path": None,
        **info,
        "evaluation": {
            "type": "gold_answer_reference",
            "method": "manual_or_llm_judge_later",
            "notes": "Curated gold answer is preserved. Current automatic scoring is intentionally lightweight.",
        },
    }


def import_cases(input_path: Path, output_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(input_path, data_only=True)
    sheet = workbook.active
    headers = [normalize_text(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    if headers[:2] != ["问题", "标准答案"]:
        raise ValueError(f"Expected first two headers to be 问题/标准答案, got {headers[:2]}")

    records: list[dict[str, Any]] = []
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        question = normalize_text(row[0] if len(row) > 0 else "")
        answer = normalize_text(row[1] if len(row) > 1 else "")
        if not question and not answer:
            continue
        if not question or not answer:
            raise ValueError(f"Row {index + 1} has missing question or answer")
        records.append(build_record(index, question, answer))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        "\n".join(json.dumps(record, ensure_ascii=False) for record in records) + "\n",
        encoding="utf-8",
    )
    return records


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=str(DEFAULT_INPUT))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()
    records = import_cases(Path(args.input), Path(args.output))
    counts: dict[str, int] = {}
    for record in records:
        counts[record["task_type"]] = counts.get(record["task_type"], 0) + 1
    print(f"Imported {len(records)} gold cases -> {args.output}")
    print(json.dumps(counts, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
