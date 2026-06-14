#!/usr/bin/env python3
"""Import reviewed badcase spreadsheet rows into TableClaw JSONL tasks."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from import_gold_cases import classify, normalize_text, parse_markdown_table


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "eval_test/test_dataset/source/300条badcase.xlsx"
DEFAULT_OUTPUT = ROOT / "eval_test/test_dataset/bad_cases.jsonl"


EXPECTED_HEADERS = {
    "问题": "question",
    "标准答案": "gold_answer",
    "模型答案": "previous_model_answer",
    "模型回复": "previous_model_response",
    "结论": "previous_label",
    "评分依据": "previous_judge_reason",
    "首Token时间(s)": "first_token_seconds",
    "外部总体耗时(s)": "external_elapsed_seconds",
    "id": "source_id",
}


def stable_id(index: int, source_id: str, question: str, answer: str) -> str:
    digest = hashlib.sha1(f"{source_id}\n{question}\n---\n{answer}".encode("utf-8")).hexdigest()[:10]
    return f"bad_case_{index:03d}_{digest}"


def number_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def badcase_tags(question: str, judge_reason: str) -> list[str]:
    text = f"{question}\n{judge_reason}"
    patterns = {
        "ranking_error": r"排名错误|排名",
        "missing_field": r"缺失|漏答|未提及|数据缺失",
        "wrong_scope": r"范围|口径|200亿省|全省|全国|市州",
        "wrong_metric": r"指标|同比|环比|占收比|金额",
        "wrong_table_or_time": r"月份|时间|表格|年份",
        "unit_error": r"单位|万元|亿元|百分比|百分点",
    }
    return [tag for tag, pattern in patterns.items() if re.search(pattern, text)]


def build_record(index: int, row: dict[str, Any], *, source_row: int) -> dict[str, Any]:
    question = normalize_text(row.get("question"))
    answer = normalize_text(row.get("gold_answer"))
    source_id = normalize_text(row.get("source_id"))
    previous_judge_reason = normalize_text(row.get("previous_judge_reason"))
    info = classify(question, answer)
    answer_table = parse_markdown_table(answer)
    return {
        "id": stable_id(index, source_id, question, answer),
        "gold_case_index": index,
        "source": {
            "raw_file": "300条badcase.xlsx",
            "sheet": "Sheet1",
            "row": source_row,
            "source_id": source_id or None,
        },
        "question": question,
        "ground_truth": answer,
        "gold_answer": answer,
        "ground_truth_table": answer_table,
        "recommended_eval_mode": "chart_data_table_only" if info["requires_visual_artifact"] else "structured_table_qa",
        "data_eval_ready": True,
        "visual_eval_ready": False,
        "retrieval_eval_ready": False,
        "table_id": None,
        "table_path": None,
        **info,
        "badcase": {
            "previous_label": normalize_text(row.get("previous_label")) or None,
            "previous_model_answer": normalize_text(row.get("previous_model_answer")) or None,
            "previous_model_response": normalize_text(row.get("previous_model_response")) or None,
            "previous_judge_reason": previous_judge_reason or None,
            "first_token_seconds": number_or_none(row.get("first_token_seconds")),
            "external_elapsed_seconds": number_or_none(row.get("external_elapsed_seconds")),
            "tags": badcase_tags(question, previous_judge_reason),
        },
        "evaluation": {
            "type": "badcase_gold_answer_reference",
            "method": "same_as_gold_cases_llm_judge",
            "notes": "Imported from reviewed badcase workbook. Prompt uses question only; gold_answer is evaluator-only.",
        },
    }


def import_cases(input_path: Path, output_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(input_path, data_only=True)
    sheet = workbook.active
    raw_headers = [normalize_text(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    headers = [EXPECTED_HEADERS.get(header, header) for header in raw_headers]
    if "question" not in headers or "gold_answer" not in headers:
        raise ValueError(f"Expected headers 问题/标准答案, got {raw_headers}")

    records: list[dict[str, Any]] = []
    for source_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {headers[idx]: values[idx] if idx < len(values) else None for idx in range(len(headers))}
        question = normalize_text(row.get("question"))
        answer = normalize_text(row.get("gold_answer"))
        if not question and not answer:
            continue
        if not question or not answer:
            raise ValueError(f"Row {source_row} has missing question or standard answer")
        records.append(build_record(len(records) + 1, row, source_row=source_row))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        "\n".join(json.dumps(record, ensure_ascii=False) for record in records) + "\n",
        encoding="utf-8",
    )
    return records


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=str(DEFAULT_INPUT))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()
    records = import_cases(Path(args.input), Path(args.output))
    counts: dict[str, int] = {}
    tag_counts: dict[str, int] = {}
    for record in records:
        counts[record["task_type"]] = counts.get(record["task_type"], 0) + 1
        for tag in record["badcase"]["tags"]:
            tag_counts[tag] = tag_counts.get(tag, 0) + 1
    print(f"Imported {len(records)} bad cases -> {args.output}")
    print(json.dumps({"task_type": counts, "badcase_tags": tag_counts}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
